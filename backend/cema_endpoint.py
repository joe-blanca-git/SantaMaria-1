
@router.post("/cema/conciliar")
async def conciliar_cema(
    cema_file: UploadFile = File(...),
    acr_file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not cema_file.filename or not acr_file.filename:
        raise HTTPException(status_code=400, detail="Arquivos inválidos")
        
    try:
        import io
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        import datetime
        import math
        
        # 1. Parse Cema File
        cema_bytes = await cema_file.read()
        try:
            df_cema = pd.read_excel(io.BytesIO(cema_bytes), header=None)
        except Exception:
            raise HTTPException(status_code=400, detail="Arquivo Cema deve ser Excel.")
            
        # Parse Dates Helper
        def parse_date(date_str):
            if not date_str:
                return None
            if isinstance(date_str, datetime.datetime):
                return date_str.date()
            if isinstance(date_str, datetime.date):
                return date_str
            date_str = str(date_str).strip()
            if hasattr(pd, 'Timestamp') and isinstance(date_str, pd.Timestamp):
                return date_str.date()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%y', '%Y/%m/%d %H:%M:%S', '%d/%m/%Y %H:%M:%S'):
                try:
                    return datetime.datetime.strptime(date_str.split()[0] if ' ' in date_str else date_str, fmt).date()
                except ValueError:
                    continue
            return None

        # Next Thursday Helper
        def get_next_thursday(d):
            if not isinstance(d, datetime.date):
                return d
            wd = d.weekday()
            if wd == 3:
                return d
            days_ahead = 3 - wd
            if days_ahead <= 0:
                days_ahead += 7
            return d + datetime.timedelta(days_ahead)

        cema_invoices = []
        for idx, row in df_cema.iterrows():
            if len(row) < 6:
                continue
            
            nf_raw = str(row[1]).strip() # Col B (index 1)
            date_raw = row[5] # Col F (index 5)
            
            if not nf_raw or nf_raw.lower() in ('nan', 'none', 'nota fiscal', 'nota', 'nota_fiscal'):
                continue
                
            nf_clean = nf_raw.split('/')[0].strip()
            
            if not nf_clean:
                continue
            
            try:
                nf_clean = str(int(float(nf_clean))).zfill(7)
            except ValueError:
                nf_clean = nf_clean.zfill(7)
            
            d_parsed = parse_date(date_raw)
            d_adjusted = get_next_thursday(d_parsed) if d_parsed else date_raw
            
            cema_invoices.append({
                'raw_nf': nf_raw,
                'nf': nf_clean,
                'prorrogacao': d_adjusted
            })
            
        if not cema_invoices:
            raise HTTPException(status_code=400, detail="Nenhuma nota fiscal encontrada no arquivo Cema.")

        # 2. Parse ACR File
        acr_bytes = await acr_file.read()
        try:
            df_acr = pd.read_excel(io.BytesIO(acr_bytes), header=None)
        except Exception:
            acr_text = acr_bytes.decode('utf-8', errors='ignore')
            sep = ';' if ';' in acr_text else ','
            df_acr = pd.read_csv(io.StringIO(acr_text), sep=sep, header=None)
            
        acr_data_by_int = {}
        for idx, row in df_acr.iterrows():
            if len(row) < 16:
                continue
            val_a = str(row[0]).strip().split('.')[0]
            if val_a != '104':
                continue
            val_b = str(row[1]).strip().upper()
            if val_b != 'DP':
                continue
                
            val_e = str(row[4]).strip().split('.')[0]
            try:
                if int(val_e) != 1:
                    continue
            except ValueError:
                continue
                
            val_d = str(row[3]).strip().split('.')[0]
            val_p = str(row[15]).strip()
            if ' ' in val_p:
                val_p = val_p.split()[0]
                
            try:
                acr_data_by_int[int(val_d)] = parse_date(val_p) or val_p
            except ValueError:
                continue
                
        # 3. Create Excel
        def format_date_to_br(d):
            if isinstance(d, datetime.date):
                return d.strftime('%d/%m/%Y')
            return str(d) if d else ""

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Prorrogações Cema"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.cell(row=1, column=1, value="Nota Fiscal")
        ws1.cell(row=1, column=2, value="Data de Prorrogação")
        
        for i, inv in enumerate(cema_invoices):
            ws1.cell(row=i+2, column=1, value=inv['nf'])
            ws1.cell(row=i+2, column=2, value=format_date_to_br(inv['prorrogacao']))
            
        ws2 = wb.create_sheet(title="Conciliação")
        ws2.views.sheetView[0].showGridLines = True
        ws2.cell(row=1, column=1, value="Nota Fiscal")
        ws2.cell(row=1, column=2, value="Data Cema (Prorrogação)")
        ws2.cell(row=1, column=3, value="Data ACR (Vencimento)")
        ws2.cell(row=1, column=4, value="Status")
        
        font_header = Font(name='Calibri', size=11, bold=True)
        font_body = Font(name='Calibri', size=11, bold=False)
        align_left = Alignment(horizontal='left', vertical='center')
        align_center = Alignment(horizontal='center', vertical='center')
        
        fill_ok = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fill_div = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        for col_idx in [1, 2]:
            cell = ws1.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left
            
        for r_idx in range(2, len(cema_invoices) + 2):
            ws1.cell(row=r_idx, column=1).font = font_body
            ws1.cell(row=r_idx, column=1).number_format = '@'
            ws1.cell(row=r_idx, column=1).alignment = align_left
            ws1.cell(row=r_idx, column=2).font = font_body
            ws1.cell(row=r_idx, column=2).alignment = align_center
            
        ws1.column_dimensions['A'].width = 16
        ws1.column_dimensions['B'].width = 24
        
        for i, inv in enumerate(cema_invoices):
            row_num = i + 2
            nf_str = inv['nf']
            d_cema = inv['prorrogacao']
            
            try:
                nf_int = int(nf_str)
            except ValueError:
                nf_int = None
                
            d_acr = None
            status = "Não encontrado no CSV"
            status_fill = None
            
            if nf_int is not None and nf_int in acr_data_by_int:
                d_acr = acr_data_by_int[nf_int]
                
                if isinstance(d_cema, datetime.date) and isinstance(d_acr, datetime.date):
                    if d_cema == d_acr:
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                else:
                    if str(d_cema).strip() == str(d_acr).strip():
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                        
            ws2.cell(row=row_num, column=1, value=nf_str)
            ws2.cell(row=row_num, column=2, value=format_date_to_br(d_cema))
            ws2.cell(row=row_num, column=3, value=format_date_to_br(d_acr))
            
            status_cell = ws2.cell(row=row_num, column=4, value=status)
            if status_fill:
                status_cell.fill = status_fill
                
        for col_idx in [1, 2, 3, 4]:
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left
            
        for r_idx in range(2, len(cema_invoices) + 2):
            ws2.cell(row=r_idx, column=1).font = font_body
            ws2.cell(row=r_idx, column=1).number_format = '@'
            ws2.cell(row=r_idx, column=1).alignment = align_left
            ws2.cell(row=r_idx, column=2).font = font_body
            ws2.cell(row=r_idx, column=2).alignment = align_center
            ws2.cell(row=r_idx, column=3).font = font_body
            ws2.cell(row=r_idx, column=3).alignment = align_center
            ws2.cell(row=r_idx, column=4).font = font_body
            ws2.cell(row=r_idx, column=4).alignment = align_left
            
        ws2.column_dimensions['A'].width = 16
        ws2.column_dimensions['B'].width = 24
        ws2.column_dimensions['C'].width = 24
        ws2.column_dimensions['D'].width = 24
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = cema_file.filename.rsplit('.', 1)[0] + "_extraido.xlsx"
        
        from app.services.importacao_service import ImportacaoService
        ImportacaoService(db).registrar_importacao(filename, "xlsx", "Prorrogação - Cema")
        
        headers_response = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_response
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
