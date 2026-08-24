import sys
with open('c:/Users/tania.canedo/Documents/git/SantaMaria/backend/app/routers/importacoes.py', 'r', encoding='utf-8') as f:
    content = f.read()

patch_code = '''
async def conciliar_composicao_ws(wb, acr_file, rows_to_export, font_header, font_body, align_center, align_left, align_right):
    if not acr_file or not acr_file.filename:
        return
    acr_bytes = await acr_file.read()
    import io
    import pandas as pd
    from openpyxl.styles import PatternFill
    try:
        df_acr = pd.read_excel(io.BytesIO(acr_bytes))
    except Exception:
        acr_text = acr_bytes.decode('utf-8', errors='ignore')
        sep = ';' if ';' in acr_text else ','
        df_acr = pd.read_csv(io.StringIO(acr_text), sep=sep)
    df_acr.columns = [str(c).strip().lower() for c in df_acr.columns]
    col_titulo = next((c for c in df_acr.columns if 'titulo' in c or 'título' in c), None)
    col_parcela = next((c for c in df_acr.columns if '/p' in c or 'parcela' in c), None)
    col_saldo = next((c for c in df_acr.columns if 'saldo' in c), None)
    acr_data = {}
    if col_titulo and col_saldo:
        for idx, row in df_acr.iterrows():
            titulo = str(row[col_titulo]).strip()
            if titulo.endswith('.0'):
                titulo = titulo[:-2]
            if col_parcela:
                parcela_val = str(row[col_parcela]).strip()
                try:
                    if int(float(parcela_val)) != 1:
                        continue
                except ValueError:
                    pass
            saldo_val = row[col_saldo]
            try:
                if isinstance(saldo_val, str):
                    saldo_val = saldo_val.replace('R$', '').replace('.', '').replace(',', '.').strip()
                saldo_float = float(saldo_val)
                acr_data[titulo] = saldo_float
            except ValueError:
                continue
    ws2 = wb.create_sheet(title="Conciliação")
    headers2 = ['Nota Fiscal', 'Parcela', 'Status NF', 'Valor Composição', 'Valor ACR', 'Status Valor']
    ws2.append(headers2)
    for col_idx, col_name in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.alignment = align_left if col_idx <= 3 else align_right
    fill_ok = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_err = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    accounting_format = '_("R$"* #,##0.00_);_("R$"* (#,##0.00);_("R$"* "-"_);_(@_)'
    for i, item in enumerate(rows_to_export):
        nf = item.get('Nota Fiscal') or item.get('nf')
        if not nf:
            continue
        nf_str = str(nf).strip()
        val_comp = item.get('Valor Liquido') or item.get('valor_liquido') or 0.0
        status_nf = 'Não Encontrado'
        status_val = '-'
        val_acr = 0.0
        nf_clean = nf_str.lstrip('0')
        found_key = None
        if nf_str in acr_data:
            found_key = nf_str
        else:
            for k in acr_data.keys():
                if k.lstrip('0') == nf_clean:
                    found_key = k
                    break
        if found_key:
            status_nf = 'Encontrado'
            val_acr = acr_data[found_key]
            if abs(val_comp - val_acr) < 0.01:
                status_val = 'OK'
            else:
                status_val = 'Divergente'
        r_idx = i + 2
        ws2.cell(row=r_idx, column=1, value=nf_str).font = font_body
        ws2.cell(row=r_idx, column=1).alignment = align_left
        ws2.cell(row=r_idx, column=2, value="01").font = font_body
        ws2.cell(row=r_idx, column=2).alignment = align_left
        ws2.cell(row=r_idx, column=3, value=status_nf).font = font_body
        ws2.cell(row=r_idx, column=3).alignment = align_left
        c4 = ws2.cell(row=r_idx, column=4, value=val_comp)
        c4.font = font_body
        c4.number_format = accounting_format
        c4.alignment = align_right
        c5 = ws2.cell(row=r_idx, column=5, value=val_acr if status_nf == 'Encontrado' else None)
        c5.font = font_body
        c5.number_format = accounting_format
        c5.alignment = align_right
        c6 = ws2.cell(row=r_idx, column=6, value=status_val)
        c6.font = font_body
        c6.alignment = align_right
        if status_nf == 'Encontrado' and status_val == 'OK':
            for c in range(1, 7):
                ws2.cell(row=r_idx, column=c).fill = fill_ok
        else:
            for c in range(1, 7):
                ws2.cell(row=r_idx, column=c).fill = fill_err
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 18
    ws2.column_dimensions['F'].width = 16

'''
content = content.replace('@router.post("/atacadao/extrair")', patch_code + '\n@router.post("/atacadao/extrair")')

# Replace extrair_atacadao signature
content = content.replace(
    'async def extrair_atacadao(file: UploadFile = File(...), db: Session = Depends(get_db)):',
    'async def extrair_atacadao(file: UploadFile = File(...), acr_file: UploadFile = File(...), db: Session = Depends(get_db)):'
)

# Insert conciliation logic before wb.save in extrair_atacadao
content = content.replace(
    '# Salvar em buffer\n        output = io.BytesIO()\n        wb.save(output)',
    'await conciliar_composicao_ws(wb, acr_file, rows_to_export, font_header, font_body, align_center, align_left, align_right)\n\n        # Salvar em buffer\n        output = io.BytesIO()\n        wb.save(output)'
)

# Replace extrair_sendas signature
content = content.replace(
    'async def extrair_sendas(file: UploadFile = File(...)):',
    'async def extrair_sendas(file: UploadFile = File(...), acr_file: UploadFile = File(...), db: Session = Depends(get_db)):'
)

# Insert conciliation logic before wb.save in extrair_sendas
content = content.replace(
    '# Save to buffer\n        output = io.BytesIO()\n        wb.save(output)',
    '# Transform notas_fiscais + abatimentos for conciliar (we need "Nota Fiscal" and "Valor Liquido")\n        all_items = []\n        for n in notas_fiscais:\n            all_items.append({"Nota Fiscal": n["Nota Fiscal"], "Valor Liquido": n["Valor Liquido"]})\n        for a in abatimentos:\n            all_items.append({"Nota Fiscal": a["Nota Fiscal"], "Valor Liquido": a["Valor Liquido"]})\n        await conciliar_composicao_ws(wb, acr_file, all_items, font_header, font_body, align_center, align_left, align_right)\n\n        # Save to buffer\n        output = io.BytesIO()\n        wb.save(output)'
)

with open('c:/Users/tania.canedo/Documents/git/SantaMaria/backend/app/routers/importacoes.py', 'w', encoding='utf-8') as f:
    f.write(content)
print("Done")
