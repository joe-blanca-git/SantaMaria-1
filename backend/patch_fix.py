import sys

with open('c:/Users/tania.canedo/Documents/git/SantaMaria/backend/app/routers/importacoes.py', 'r', encoding='utf-8') as f:
    content = f.read()

# Extract the existing conciliar_composicao_ws block to replace it
import re
pattern = re.compile(r'async def conciliar_composicao_ws.*?ws2\.column_dimensions\[\'F\'\]\.width = 16\n', re.DOTALL)

new_func = '''async def conciliar_composicao_ws(wb, acr_file, rows_to_export, font_header, font_body, align_center, align_left, align_right):
    if not acr_file or not acr_file.filename:
        return
    acr_bytes = await acr_file.read()
    import io
    import pandas as pd
    from openpyxl.styles import PatternFill
    try:
        df_acr = pd.read_excel(io.BytesIO(acr_bytes), header=None)
    except Exception:
        acr_text = acr_bytes.decode('utf-8', errors='ignore')
        sep = ';' if ';' in acr_text else ','
        df_acr = pd.read_csv(io.StringIO(acr_text), sep=sep, header=None)
        
    # Find column indices dynamically by scanning the first 30 rows
    col_titulo = -1
    col_parcela = -1
    col_saldo = -1
    for idx, row in df_acr.head(30).iterrows():
        row_strs = [str(x).strip().lower() for x in row.values]
        for c_idx, val in enumerate(row_strs):
            if 'título' in val or 'titulo' in val:
                col_titulo = c_idx
            if '/p' in val or 'parcela' in val:
                col_parcela = c_idx
            if 'saldo' in val:
                col_saldo = c_idx
        if col_titulo != -1 and col_saldo != -1:
            break
            
    # Fallbacks baseados na prorrogação
    if col_titulo == -1:
        col_titulo = 3
    if col_parcela == -1:
        col_parcela = 4
    if col_saldo == -1:
        col_saldo = 17 # typical if 15 is due date? we'll try to rely on dynamic search primarily.
        
    acr_data = {}
    for idx, row in df_acr.iterrows():
        if len(row) <= max(col_titulo, col_saldo, col_parcela):
            continue
            
        # Filtros rigorosos iguais aos da prorrogação
        val_a = str(row[0]).strip().split('.')[0]
        if val_a != '104':
            continue
        val_b = str(row[1]).strip().upper()
        if val_b != 'DP':
            continue
            
        titulo = str(row[col_titulo]).strip()
        if titulo.endswith('.0'):
            titulo = titulo[:-2]
            
        if col_parcela != -1:
            parcela_val = str(row[col_parcela]).strip()
            try:
                if int(float(parcela_val)) != 1:
                    continue
            except ValueError:
                pass
                
        saldo_val = row[col_saldo]
        try:
            if isinstance(saldo_val, str):
                saldo_val = saldo_val.replace('R$', '').replace('.', '').replace(',', '.').strip()
            saldo_float = float(saldo_val)
            acr_data[titulo] = saldo_float
        except ValueError:
            continue
            
    ws2 = wb.create_sheet(title="Conciliação")
    headers2 = ['Nota Fiscal', 'Parcela', 'Status NF', 'Valor Composição', 'Valor ACR', 'Status Valor']
    ws2.append(headers2)
    for col_idx, col_name in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.alignment = align_left if col_idx <= 3 else align_right
    fill_ok = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_err = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    accounting_format = '_("R$"* #,##0.00_);_("R$"* (#,##0.00);_("R$"* "-"_);_(@_)'
    
    for i, item in enumerate(rows_to_export):
        nf = item.get('Nota Fiscal') or item.get('nf')
        if not nf:
            continue
        nf_str = str(nf).strip()
        val_comp = item.get('Valor Liquido') or item.get('valor_liquido') or 0.0
        status_nf = 'Não Encontrado'
        status_val = '-'
        val_acr = 0.0
        nf_clean = nf_str.lstrip('0')
        found_key = None
        if nf_str in acr_data:
            found_key = nf_str
        else:
            for k in acr_data.keys():
                if k.lstrip('0') == nf_clean:
                    found_key = k
                    break
        if found_key:
            status_nf = 'Encontrado'
            val_acr = acr_data[found_key]
            if abs(val_comp - val_acr) < 0.01:
                status_val = 'OK'
            else:
                status_val = 'Divergente'
        r_idx = i + 2
        ws2.cell(row=r_idx, column=1, value=nf_str).font = font_body
        ws2.cell(row=r_idx, column=1).alignment = align_left
        ws2.cell(row=r_idx, column=2, value="01").font = font_body
        ws2.cell(row=r_idx, column=2).alignment = align_left
        ws2.cell(row=r_idx, column=3, value=status_nf).font = font_body
        ws2.cell(row=r_idx, column=3).alignment = align_left
        c4 = ws2.cell(row=r_idx, column=4, value=val_comp)
        c4.font = font_body
        c4.number_format = accounting_format
        c4.alignment = align_right
        c5 = ws2.cell(row=r_idx, column=5, value=val_acr if status_nf == 'Encontrado' else None)
        c5.font = font_body
        c5.number_format = accounting_format
        c5.alignment = align_right
        c6 = ws2.cell(row=r_idx, column=6, value=status_val)
        c6.font = font_body
        c6.alignment = align_right
        if status_nf == 'Encontrado' and status_val == 'OK':
            for c in range(1, 7):
                ws2.cell(row=r_idx, column=c).fill = fill_ok
        else:
            for c in range(1, 7):
                ws2.cell(row=r_idx, column=c).fill = fill_err
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 18
    ws2.column_dimensions['F'].width = 16
'''

content = pattern.sub(new_func, content)
with open('c:/Users/tania.canedo/Documents/git/SantaMaria/backend/app/routers/importacoes.py', 'w', encoding='utf-8') as f:
    f.write(content)
print("Updated conciliar_composicao_ws")
