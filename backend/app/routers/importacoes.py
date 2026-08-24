from fastapi import APIRouter, Depends, Query, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import io
import pandas as pd
from app.core.database import get_db
from app.schemas.importacao import ImportacaoPaginatedResponse
from app.services.importacao_service import ImportacaoService
from app.services.ia_service import IAService
from app.services.dashboard_service import DashboardService
from app.repositories.categoria_repository import CategoriaRepository
from app.repositories.colaborador_repository import ColaboradorRepository
from pydantic import BaseModel
from typing import List, Optional

router = APIRouter()

def get_service(db: Session = Depends(get_db)):
    return ImportacaoService(db)

@router.get("/", response_model=ImportacaoPaginatedResponse)
def get_importacoes(
    page: int = Query(1, ge=1),
    size: int = Query(10, ge=1, le=100),
    search: str = Query(None, description="Busca por nome de arquivo ou tipo"),
    categoria: str = Query(None, description="Categoria exata (ex: Composição, Prorrogação)"),
    service: ImportacaoService = Depends(get_service)
):
    return service.listar_importacoes(page=page, size=size, search=search, categoria=categoria)

@router.post("/ia/analise-extrato")
async def analise_extrato_ia(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    empresa_nome: str = Form(...),
    db: Session = Depends(get_db)
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo inválido")
        
    try:
        t0_overall = time.time() if 'time' in globals() else __import__('time').time()
        content = await file.read()
        
        # Buscar listas base do DB para passar como contexto
        cat_repo = CategoriaRepository(db)
        colab_repo = ColaboradorRepository(db)
        
        categorias_db, _ = cat_repo.get_all(limit=1000)
        nomes_categorias = [c.nome for c in categorias_db]
        
        colabs_db, _ = colab_repo.get_all(limit=5000)
        nomes_colaboradores = [c.nome for c in colabs_db]
        
        ia = IAService()
        ia_result = await ia.analisar_extrato(
            file_content=content,
            file_name=file.filename,
            categorias=nomes_categorias,
            colaboradores=nomes_colaboradores,
            empresa_context=empresa_nome
        )
        
        despesas_brutas = ia_result.get("despesas", [])
        metrics = ia_result.get("metrics", {})
        file_to_delete = ia_result.get("file_name_to_delete")
        
        # Se for necessário deletar arquivo da Files API, agenda em background
        if file_to_delete:
            background_tasks.add_task(ia.deletar_arquivo, file_to_delete)
            
        # Consolidar os valores por Colaborador + Categoria (somar)
        t0_consolidation = (__import__('time').time() if 'time' not in globals() else time.time())
        consolidadas = {}
        for d in despesas_brutas:
            chave = f"{d['colaborador']}|{d['categoria']}"
            if chave not in consolidadas:
                consolidadas[chave] = d
            else:
                consolidadas[chave]['valor'] += d['valor']
                
        resultado_final = list(consolidadas.values())
        
        # Atualiza métricas
        post_proc_ms = metrics.get("post_processing_ms", 0.0)
        time_module = (__import__('time') if 'time' not in globals() else time)
        metrics["post_processing_ms"] = round(post_proc_ms + (time_module.time() - t0_consolidation) * 1000, 2)
        metrics["total_ms"] = round((time_module.time() - t0_overall) * 1000, 2)
        
        # Logging exigido de métricas de latência
        print(f"[IA] Upload PDF: {metrics['upload_pdf_ms']} ms")
        print(f"[IA] File ready: {metrics['file_ready_ms']} ms")
        print(f"[IA] Gemini generation: {metrics['gemini_generation_ms']} ms")
        print(f"[IA] Structured output: {metrics['structured_output_ms']} ms")
        print(f"[IA] Post processing: {metrics['post_processing_ms']} ms")
        print(f"[IA] Total: {metrics['total_ms']} ms")
        
        return {"sucesso": True, "dados": resultado_final}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

from app.schemas.movimentacao import SalvarImportacaoIAPayload
from app.services.movimentacao_service import MovimentacaoService

@router.post("/ia/salvar")
def salvar_movimentacoes_ia(
    payload: SalvarImportacaoIAPayload,
    db: Session = Depends(get_db)
):
    try:
        service = MovimentacaoService(db)
        resultado = service.salvar_importacao_ia(payload)
        return resultado
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/{id_importacao}")
def excluir_importacao(
    id_importacao: int,
    db: Session = Depends(get_db)
):
    try:
        service = ImportacaoService(db)
        sucesso = service.excluir_importacao(id_importacao)
        if not sucesso:
            raise HTTPException(status_code=404, detail="Importação não encontrada")
        return {"sucesso": True, "mensagem": "Importação e movimentações vinculadas excluídas com sucesso"}
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/dashboard")
def get_dashboard(
    data_inicio: str = Query(None),
    data_fim: str = Query(None),
    id_empresa: int = Query(None),
    id_colaborador: int = Query(None),
    id_categoria: int = Query(None),
    db: Session = Depends(get_db)
):
    try:
        service = DashboardService(db)
        return service.obter_dados(
            data_inicio=data_inicio,
            data_fim=data_fim,
            id_empresa=id_empresa,
            id_colaborador=id_colaborador,
            id_categoria=id_categoria
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/dashboard/analitico")
def get_dashboard_analitico(
    data_inicio: str = Query(None),
    data_fim: str = Query(None),
    id_empresa: int = Query(None),
    id_colaborador: int = Query(None),
    id_categoria: int = Query(None),
    db: Session = Depends(get_db)
):
    try:
        service = DashboardService(db)
        return service.obter_dados_analitico(
            data_inicio=data_inicio,
            data_fim=data_fim,
            id_empresa=id_empresa,
            id_colaborador=id_colaborador,
            id_categoria=id_categoria
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

import re
import io
import pandas as pd
from fastapi.responses import StreamingResponse
from html.parser import HTMLParser

class AtacadaoHTMLParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.tables = []
        self.current_table = []
        self.current_row = []
        self.current_cell = []
        self.in_table = False
        self.in_tr = False
        self.in_td_or_th = False

    def handle_starttag(self, tag, attrs):
        if tag == 'table':
            self.in_table = True
            self.current_table = []
        elif tag == 'tr' and self.in_table:
            self.in_tr = True
            self.current_row = []
        elif tag in ('td', 'th') and self.in_tr:
            self.in_td_or_th = True
            self.current_cell = []

    def handle_endtag(self, tag):
        if tag == 'table':
            self.in_table = False
            if self.current_table:
                self.tables.append(self.current_table)
        elif tag == 'tr' and self.in_table:
            self.in_tr = False
            if self.current_row:
                self.current_table.append(self.current_row)
        elif tag in ('td', 'th') and self.in_tr:
            self.in_td_or_th = False
            cell_text = "".join(self.current_cell).strip()
            self.current_row.append(cell_text)

    def handle_data(self, data):
        if self.in_td_or_th:
            self.current_cell.append(data)


async def conciliar_composicao_ws(wb, acr_file, rows_to_export, font_header, font_body, align_center, align_left, align_right):
    if not acr_file or not acr_file.filename:
        return
    acr_bytes = await acr_file.read()
    import io
    import pandas as pd
    from openpyxl.styles import PatternFill
    import logging
    debug_logger = logging.getLogger("composicao_debug")
    debug_logger.setLevel(logging.DEBUG)
    # Clear handlers
    debug_logger.handlers = []
    fh = logging.FileHandler('c:/Users/tania.canedo/Documents/git/SantaMaria/backend/app/routers/import_debug.log', mode='w', encoding='utf-8')
    fh.setFormatter(logging.Formatter('%(asctime)s - %(message)s'))
    debug_logger.addHandler(fh)
    
    debug_logger.info("--- INICIANDO CONCILIACAO COMPOSICAO ---")
    debug_logger.info(f"acr_file.filename: {acr_file.filename}")
    debug_logger.info(f"rows_to_export count: {len(rows_to_export)}")
    if rows_to_export:
        debug_logger.info(f"Exemplo item export: {rows_to_export[0]}")
        
    try:
        with open('c:/Users/tania.canedo/Documents/git/SantaMaria/backend/ACR_debug.xlsx', 'wb') as debug_f:
            debug_f.write(acr_bytes)
        debug_logger.info("Salvo ACR_debug.xlsx com sucesso")
    except Exception as ex_save:
        debug_logger.error(f"Erro ao salvar ACR_debug.xlsx: {str(ex_save)}")
        
    try:
        df_acr = pd.read_excel(io.BytesIO(acr_bytes), header=None)
        debug_logger.info("Lido com read_excel")
    except Exception as ex:
        debug_logger.error(f"Erro read_excel: {str(ex)}")
        acr_text = acr_bytes.decode('utf-8', errors='ignore')
        sep = ';' if ';' in acr_text else ','
        df_acr = pd.read_csv(io.StringIO(acr_text), sep=sep, header=None)
        debug_logger.info(f"Lido com read_csv, sep={sep}")
        
    debug_logger.info(f"df_acr shape: {df_acr.shape}")
    debug_logger.info(f"Primeiras 10 linhas do df_acr:\n{df_acr.head(10).to_string()}")
        
    col_titulo = -1
    col_parcela = -1
    col_saldo = -1
    
    # 1. Busca pelo nome no cabeçalho
    for idx, row in df_acr.head(30).iterrows():
        row_strs = [str(x).strip().lower() for x in row.values]
        for c_idx, val in enumerate(row_strs):
            if val == 'titulo' or val == 'título' or 'titulo' in val or 'título' in val:
                if col_titulo == -1: col_titulo = c_idx
            if val == '/p' or val == 'parcela' or '/p' in val or 'parcela' in val:
                if col_parcela == -1: col_parcela = c_idx
            if val == 'saldo' or 'saldo' in val or 'valor l' in val:
                if col_saldo == -1: col_saldo = c_idx
        if col_titulo != -1 and col_parcela != -1 and col_saldo != -1:
            break
            
    debug_logger.info(f"Apos busca de cabeçalho: col_titulo={col_titulo}, col_parcela={col_parcela}, col_saldo={col_saldo}")
            
    # 2. Heurística Robusta: auto-detecção cruzando NFs e valores
    comp_map = {}
    for item in rows_to_export:
        nf = str(item.get('Nota Fiscal') or item.get('nf')).strip()
        if nf:
            v = abs(item.get('Valor Liquido') or item.get('valor_liquido') or 0.0)
            if v > 0:
                comp_map[nf.lstrip('0')] = v
                
    debug_logger.info(f"comp_map (NFs da Composicao lstrip): {list(comp_map.keys())[:10]}")
    debug_logger.info(f"comp_map values: {list(comp_map.values())[:10]}")
                
    possible_titulo_cols = {}
    possible_saldo_cols = {}
    
    for idx, row in df_acr.iterrows():
        row_vals = list(row.values)
        for c_idx, val in enumerate(row_vals):
            if pd.isna(val):
                continue
            val_str = str(val).strip().split('.')[0]
            if val_str.endswith('.0'): val_str = val_str[:-2]
            val_clean = val_str.lstrip('0')
            
            if val_clean in comp_map:
                possible_titulo_cols[c_idx] = possible_titulo_cols.get(c_idx, 0) + 1
                expected_val = comp_map[val_clean]
                for val_c_idx, val_cell in enumerate(row_vals):
                    try:
                        if isinstance(val_cell, str):
                            val_cell_clean = val_cell.replace('R$', '').replace('.', '').replace(',', '.').strip()
                            val_cell_float = abs(float(val_cell_clean))
                        else:
                            val_cell_float = abs(float(val_cell))
                        if val_cell_float > 0 and abs(val_cell_float - expected_val) < 0.01:
                            possible_saldo_cols[val_c_idx] = possible_saldo_cols.get(val_c_idx, 0) + 1
                    except Exception:
                        pass
                        
    debug_logger.info(f"possible_titulo_cols: {possible_titulo_cols}")
    debug_logger.info(f"possible_saldo_cols: {possible_saldo_cols}")
                        
    if col_titulo == -1 and possible_titulo_cols:
        col_titulo = max(possible_titulo_cols, key=possible_titulo_cols.get)
    if col_saldo == -1 and possible_saldo_cols:
        col_saldo = max(possible_saldo_cols, key=possible_saldo_cols.get)
        
    # Fallbacks finais baseados nos formatos mais comuns
    if col_titulo == -1:
        col_titulo = 2 # Coluna C
    if col_parcela == -1:
        if col_titulo != -1 and col_titulo + 1 < df_acr.shape[1]:
            col_parcela = col_titulo + 1
        else:
            col_parcela = 3 # Coluna D
    if col_saldo == -1:
        col_saldo = 21 # Coluna V
        
    debug_logger.info(f"FINAL col_titulo={col_titulo}, col_parcela={col_parcela}, col_saldo={col_saldo}")

    acr_data = {}
    for idx, row in df_acr.iterrows():
        if len(row) <= col_titulo or len(row) <= col_saldo:
            continue
            
        val_d_raw = row[col_titulo]
        if pd.isna(val_d_raw):
            continue
            
        titulo = str(val_d_raw).strip().split('.')[0]
        if not any(char.isdigit() for char in titulo):
            continue
            
        if titulo.endswith('.0'):
            titulo = titulo[:-2]
            
        # Pular se o titulo lstrip for vazio ou não numérico
        titulo_clean = titulo.lstrip('0')
        if not titulo_clean:
            continue
            
        if col_parcela < len(row):
            parcela_val = str(row[col_parcela]).strip().split('.')[0]
            try:
                # Filtrar apenas Parcela 01 (igual nas Prorrogações)
                if int(parcela_val) != 1:
                    continue
            except ValueError:
                pass
                
        saldo_val = row[col_saldo]
        try:
            if isinstance(saldo_val, str):
                saldo_val = saldo_val.replace('R$', '').replace('.', '').replace(',', '.').strip()
            saldo_float = float(saldo_val)
            acr_data[titulo_clean] = abs(saldo_float)
        except ValueError:
            continue
            
    debug_logger.info(f"acr_data count: {len(acr_data)}")
    debug_logger.info(f"Exemplo acr_data: {list(acr_data.items())[:10]}")
            
    ws2 = wb.create_sheet(title="Conciliação")
    headers2 = ['Nota Fiscal', 'Parcela', 'Status NF', 'Valor Composição', 'Valor ACR', 'Status Valor', 'Diferença']
    ws2.append(headers2)
    for col_idx, col_name in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.alignment = align_left if col_idx <= 3 else align_right
    fill_ok = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_err = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    accounting_format = '_("R$"* #,##0.00_);_("R$"* (#,##0.00);_("R$"* "-"_);_(@_)'
    
    for i, item in enumerate(rows_to_export):
        nf = item.get('Nota Fiscal') or item.get('nf')
        if not nf:
            continue
        nf_str = str(nf).strip()
        val_comp = item.get('Valor Liquido') or item.get('valor_liquido') or 0.0
        
        status_nf = 'Não Encontrado'
        status_val = '-'
        val_acr = 0.0
        
        nf_clean = nf_str.lstrip('0')
        if nf_clean.endswith('T') or nf_clean.endswith('t'):
            nf_base = nf_clean[:-1]
        else:
            nf_base = nf_clean
            
        found_key = None
        if nf_clean in acr_data:
            found_key = nf_clean
        elif nf_base in acr_data:
            found_key = nf_base
                    
        if found_key:
            status_nf = 'Encontrado'
            val_acr = acr_data[found_key]
            if abs(abs(val_comp) - val_acr) < 0.01:
                status_val = 'OK'
            else:
                status_val = 'Divergente'
                
        r_idx = i + 2
        ws2.cell(row=r_idx, column=1, value=nf_str).font = font_body
        ws2.cell(row=r_idx, column=1).alignment = align_left
        ws2.cell(row=r_idx, column=2, value="01").font = font_body
        ws2.cell(row=r_idx, column=2).alignment = align_left
        ws2.cell(row=r_idx, column=3, value=status_nf).font = font_body
        ws2.cell(row=r_idx, column=3).alignment = align_left
        c4 = ws2.cell(row=r_idx, column=4, value=val_comp)
        c4.font = font_body
        c4.number_format = accounting_format
        c4.alignment = align_right
        c5 = ws2.cell(row=r_idx, column=5, value=val_acr if status_nf == 'Encontrado' else None)
        c5.font = font_body
        c5.number_format = accounting_format
        c5.alignment = align_right
        c6 = ws2.cell(row=r_idx, column=6, value=status_val)
        c6.font = font_body
        c6.alignment = align_right
        diferenca_val = (val_comp - val_acr) if status_val == 'Divergente' else None
        c7 = ws2.cell(row=r_idx, column=7, value=diferenca_val)
        c7.font = font_body
        c7.number_format = accounting_format
        c7.alignment = align_right
        if status_nf == 'Encontrado' and status_val == 'OK':
            for c in range(1, 8):
                ws2.cell(row=r_idx, column=c).fill = fill_ok
        else:
            for c in range(1, 8):
                ws2.cell(row=r_idx, column=c).fill = fill_err
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 18
    ws2.column_dimensions['F'].width = 16
    ws2.column_dimensions['G'].width = 18

@router.post("/atacadao/extrair")
async def extrair_atacadao(file: UploadFile = File(...), acr_file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo inválido")
        
    try:
        content_bytes = await file.read()
        content = content_bytes.decode("utf-8", errors="ignore")
        
        parser = AtacadaoHTMLParser()
        parser.feed(content)
        
        def parse_float(val_str):
            if not val_str:
                return 0.0
            cleaned = val_str.replace('.', '').replace(',', '.').strip()
            try:
                return float(cleaned)
            except ValueError:
                return 0.0

        # Extrair o Total do Depósito do texto do HTML
        total_deposito = 0.0
        match_total = re.search(r'Total do Dep&oacute;sito:\s*([\d\.,]+)', content, re.IGNORECASE)
        if not match_total:
            match_total = re.search(r'Total do Depósito:\s*([\d\.,]+)', content, re.IGNORECASE)
        if match_total:
            total_deposito = parse_float(match_total.group(1))

        rows_to_export = []
        for table in parser.tables:
            if not table:
                continue
            headers = [h.strip() for h in table[0]]
            if 'N Fiscal' not in headers:
                continue
            
            n_fiscal_idx = headers.index('N Fiscal')
            val_idx = headers.index('Valor') if 'Valor' in headers else None
            abat_idx = headers.index('Abatimento') if 'Abatimento' in headers else None
            liq_idx = headers.index('Vlr liquido') if 'Vlr liquido' in headers else None
            
            for row in table[1:]:
                if len(row) <= n_fiscal_idx:
                    continue
                
                n_fiscal_raw = row[n_fiscal_idx].strip()
                if not n_fiscal_raw or n_fiscal_raw.lower() == 'total' or 'total' in n_fiscal_raw.lower():
                    continue
                
                # Filtro: Apenas notas fiscais que comecem com '0000'
                if not n_fiscal_raw.startswith('0000'):
                    continue
                
                # Formatação: Remover os dois primeiros zeros (de 4 zeros para 2 zeros)
                n_fiscal_formatted = n_fiscal_raw[2:]
                
                # Parcela:
                # Se terminar com 'T', Parcela '02', senão '01'
                parcela = '02' if n_fiscal_raw.endswith('T') else '01'
                
                valor_total = parse_float(row[val_idx]) if val_idx is not None and val_idx < len(row) else 0.0
                abatimento = parse_float(row[abat_idx]) if abat_idx is not None and abat_idx < len(row) else 0.0
                
                if liq_idx is not None and liq_idx < len(row):
                    valor_liquido = parse_float(row[liq_idx])
                else:
                    valor_liquido = valor_total
                    
                rows_to_export.append({
                    'Nota Fiscal': n_fiscal_formatted,
                    'Parcela': parcela,
                    'Abatimento': abatimento if abatimento != 0.0 else None,
                    'Valor Liquido': valor_liquido,
                    'Valor Total': valor_total
                })
        
        if not rows_to_export:
            raise HTTPException(status_code=400, detail="Nenhum dado de Nota Fiscal começando com '0000' foi encontrado no HTML.")
            
        # Caso não consiga extrair o total do depósito via texto, calcula a soma do valor líquido
        if total_deposito == 0.0:
            total_deposito = sum((r['Valor Liquido'] or 0.0) for r in rows_to_export)

        # Geração do arquivo Excel formatado com openpyxl
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Composição de Pagamento"
        
        # Ativar linhas de grade
        ws.views.sheetView[0].showGridLines = True

        # Cabeçalhos exatamente conforme o print
        headers = ['Nota Fiscal', 'Parcela', 'Abatimento', 'Valor Líquido', '', '', 'Valor Total']
        ws.append(headers)

        # Preenchimento das linhas
        for i, item in enumerate(rows_to_export):
            # Apenas a linha 2 da coluna G possui o total geral
            val_total = total_deposito if i == 0 else None
            
            ws.append([
                item['Nota Fiscal'],
                item['Parcela'],
                item['Abatimento'],
                item['Valor Liquido'],
                '',
                '',
                val_total
            ])

        # Definição das fontes e alinhamento
        font_header = Font(name='Calibri', size=11, bold=True)
        font_body = Font(name='Calibri', size=11, bold=False)
        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')

        # Formato contábil do Excel (R$ e valor alinhados nas extremidades da célula)
        accounting_format = '_("R$"* #,##0.00_);_("R$"* (#,##0.00);_("R$"* "-"_);_(@_)'

        # Estilizar cabeçalho
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            if col_idx in [1, 2]:
                cell.alignment = align_left
            elif col_idx in [3, 4, 7]:
                cell.alignment = align_right

        # Estilizar o corpo e aplicar formatação de números
        for r_idx in range(2, len(rows_to_export) + 2):
            # Nota Fiscal (Col 1)
            cell_nf = ws.cell(row=r_idx, column=1)
            cell_nf.font = font_body
            cell_nf.number_format = '@' # Formato Texto
            cell_nf.alignment = align_left
            
            # Parcela (Col 2)
            cell_par = ws.cell(row=r_idx, column=2)
            cell_par.font = font_body
            cell_par.number_format = '@' # Formato Texto
            cell_par.alignment = align_left
            
            # Abatimento (Col 3)
            cell_ab = ws.cell(row=r_idx, column=3)
            cell_ab.font = font_body
            cell_ab.number_format = accounting_format
            cell_ab.alignment = align_right
            
            # Valor Líquido (Col 4)
            cell_liq = ws.cell(row=r_idx, column=4)
            cell_liq.font = font_body
            cell_liq.number_format = accounting_format
            cell_liq.alignment = align_right
            
            # Valor Total (Col 7)
            cell_tot = ws.cell(row=r_idx, column=7)
            cell_tot.font = font_body
            if cell_tot.value is not None:
                cell_tot.number_format = accounting_format
                cell_tot.alignment = align_right

        # Definir larguras de coluna estáticas para se adequar ao print
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 5
        ws.column_dimensions['F'].width = 5
        ws.column_dimensions['G'].width = 20

        await conciliar_composicao_ws(wb, acr_file, rows_to_export, font_header, font_body, align_center, align_left, align_right)

        # Salvar em buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = file.filename.rsplit('.', 1)[0] + "_extraido.xlsx"
        
        # Registrar no banco
        ImportacaoService(db).registrar_importacao(filename, "xlsx", "Composição - Atacadão")
        
        headers_response = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_response
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/sendas/extrair")
async def extrair_sendas(file: UploadFile = File(...), acr_file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo inválido")
        
    try:
        content_bytes = await file.read()
        
        # Load the spreadsheet
        df = pd.read_excel(io.BytesIO(content_bytes), header=None)
        
        # Resolve H and X column names/indices (H is 8th column -> index 7, X is 24th column -> index 23)
        if df.shape[1] < 24:
            raise HTTPException(
                status_code=400, 
                detail=f"A planilha importada precisa ter pelo menos 24 colunas (até a coluna X). Colunas encontradas: {df.shape[1]}."
            )
            
        h_col_idx = 7
        x_col_idx = 23
        
        def parse_float(val):
            if pd.isna(val):
                return 0.0
            if isinstance(val, (int, float)):
                return float(val)
            cleaned = str(val).replace('.', '').replace(',', '.').replace('R$', '').strip()
            try:
                return float(cleaned)
            except ValueError:
                return 0.0

        notas_fiscais = []
        devolucoes = []
        abatimentos = []

        # Iterate rows starting from index 0 or 1 depending on header detection
        for idx, row in df.iterrows():
            val_h_raw = row[h_col_idx]
            val_x_raw = row[x_col_idx]
            
            if pd.isna(val_h_raw):
                continue
                
            val_h = str(val_h_raw).strip()
            # Remove decimal part .0 if auto-formatted as float
            if val_h.endswith('.0'):
                val_h = val_h[:-2]
                
            val_x = parse_float(val_x_raw)
            
            # Filtro: Ignorar linhas e dados onde a coluna X estiver com 0
            if val_x == 0.0:
                continue

            # Skip header rows
            if val_h.lower() in ('h', 'código', 'nota fiscal', 'portador', 'n fiscal', 'tipo'):
                continue

            # Apply rules:
            # 1. Nota Fiscal: starts with 1 or 2 AND ends with 2
            if val_h.startswith(('1', '2')) and val_h.endswith('2'):
                # Formatação: Remover espaços e o último caractere '2', e preencher com zeros à esquerda até obter 7 dígitos
                cleaned_h = val_h.replace(' ', '')
                if cleaned_h.endswith('2'):
                    cleaned_h = cleaned_h[:-1]
                formatted_nf = cleaned_h.zfill(7)
                notas_fiscais.append((formatted_nf, val_x))
            # 2. Abatimento: starts with "AC"
            elif val_h.startswith('AC'):
                abatimentos.append((val_h, val_x))
            # 3. Devolução: starts with any number AND ends with "300"
            elif val_h.endswith('300') and val_h[0].isdigit():
                devolucoes.append((val_h, val_x))
                
        if not notas_fiscais and not devolucoes and not abatimentos:
            raise HTTPException(
                status_code=400, 
                detail="Nenhum registro correspondente aos filtros de Nota Fiscal, Devolução ou Abatimento foi encontrado na coluna H."
            )

        # Generate excel with openpyxl to match the layout
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Filtro Sendas"
        ws.views.sheetView[0].showGridLines = True

        # Headers exactly as requested
        # Col A: Nota fiscal, Col B: Valor Total
        # Col D: Valor Total, Col E: Devolução
        # Col I: Abatimento, Col J: Valor Total
        # Col L: Soma Total
        ws.cell(row=1, column=1, value="Nota fiscal")
        ws.cell(row=1, column=2, value="Valor Total")
        ws.cell(row=1, column=4, value="Valor Total")
        ws.cell(row=1, column=5, value="Devolução")
        ws.cell(row=1, column=9, value="Abatimento")
        ws.cell(row=1, column=10, value="Valor Total")
        ws.cell(row=1, column=12, value="Soma Total")

        max_rows = max(len(notas_fiscais), len(devolucoes), len(abatimentos))

        for r_idx in range(max_rows):
            row_num = r_idx + 2
            
            # 1. Nota Fiscal (Col A, B)
            if r_idx < len(notas_fiscais):
                nf_code, nf_val = notas_fiscais[r_idx]
                ws.cell(row=row_num, column=1, value=nf_code)
                ws.cell(row=row_num, column=2, value=nf_val)
                
            # 2. Devolução (Col E, D)
            if r_idx < len(devolucoes):
                dev_code, dev_val = devolucoes[r_idx]
                ws.cell(row=row_num, column=5, value=dev_code)
                ws.cell(row=row_num, column=4, value=dev_val)
                
            # 3. Abatimento (Col I, J)
            if r_idx < len(abatimentos):
                ab_code, ab_val = abatimentos[r_idx]
                ws.cell(row=row_num, column=9, value=ab_code)
                ws.cell(row=row_num, column=10, value=ab_val)

        # Write Soma total on Row 2 of Column L
        soma_nf = sum(v for _, v in notas_fiscais)
        soma_dev = sum(v for _, v in devolucoes)
        soma_abat = sum(v for _, v in abatimentos)
        soma_total = soma_nf + soma_dev + soma_abat

        ws.cell(row=2, column=12, value=soma_total)

        # Formatting
        font_header = Font(name='Calibri', size=11, bold=True)
        font_body = Font(name='Calibri', size=11, bold=False)
        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        accounting_format = '_("R$"* #,##0.00_);_("R$"* (#,##0.00);_("R$"* "-"_);_(@_)'

        # Format Headers
        active_cols = [1, 2, 4, 5, 9, 10, 12]
        for col_idx in active_cols:
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            if col_idx in [1, 5, 9]:
                cell.alignment = align_left
            elif col_idx in [2, 4, 10, 12]:
                cell.alignment = align_right

        # Format Body
        for r_idx in range(2, max_rows + 2):
            # Nota Fiscal (Col 1)
            cell_nf = ws.cell(row=r_idx, column=1)
            cell_nf.font = font_body
            cell_nf.number_format = '@'
            cell_nf.alignment = align_left
            
            # NF Valor (Col 2)
            c_nf_val = ws.cell(row=r_idx, column=2)
            c_nf_val.font = font_body
            if c_nf_val.value is not None:
                c_nf_val.number_format = accounting_format
                c_nf_val.alignment = align_right
                
            # Devolução Valor (Col 4)
            c_dev_val = ws.cell(row=r_idx, column=4)
            c_dev_val.font = font_body
            if c_dev_val.value is not None:
                c_dev_val.number_format = accounting_format
                c_dev_val.alignment = align_right
                
            # Devolução Código (Col 5)
            cell_dev = ws.cell(row=r_idx, column=5)
            cell_dev.font = font_body
            cell_dev.number_format = '@'
            cell_dev.alignment = align_left
            
            # Abatimento Código (Col 9)
            cell_ab = ws.cell(row=r_idx, column=9)
            cell_ab.font = font_body
            cell_ab.number_format = '@'
            cell_ab.alignment = align_left
            
            # Abatimento Valor (Col 10)
            c_ab_val = ws.cell(row=r_idx, column=10)
            c_ab_val.font = font_body
            if c_ab_val.value is not None:
                c_ab_val.number_format = accounting_format
                c_ab_val.alignment = align_right

        # Format Soma total (Col 12, Row 2)
        c_soma = ws.cell(row=2, column=12)
        c_soma.font = font_body
        c_soma.number_format = accounting_format
        c_soma.alignment = align_right

        # Column widths
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 5
        ws.column_dimensions['G'].width = 5
        ws.column_dimensions['H'].width = 5
        ws.column_dimensions['I'].width = 16
        ws.column_dimensions['J'].width = 16
        ws.column_dimensions['K'].width = 5
        ws.column_dimensions['L'].width = 20

        # Transform notas_fiscais + abatimentos for conciliar (we need "Nota Fiscal" and "Valor Liquido")
        all_items = []
        for n in notas_fiscais:
            all_items.append({"Nota Fiscal": n[0], "Valor Liquido": n[1]})
        for a in abatimentos:
            all_items.append({"Nota Fiscal": a[0], "Valor Liquido": a[1]})
        await conciliar_composicao_ws(wb, acr_file, all_items, font_header, font_body, align_center, align_left, align_right)

        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = file.filename.rsplit('.', 1)[0] + "_extraido.xlsx"
        
        # Registrar no banco
        ImportacaoService(db).registrar_importacao(filename, "xlsx", "Composição - Sendas")
        
        headers_response = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_response
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/atacadao/conciliar")
async def conciliar_atacadao(
    html_file: UploadFile = File(...),
    csv_file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not html_file.filename or not csv_file.filename:
        raise HTTPException(status_code=400, detail="Arquivos inválidos")
        
    try:
        # 1. Parse HTML file
        html_bytes = await html_file.read()
        html_content = html_bytes.decode('utf-8', errors='ignore')
        
        parser = AtacadaoHTMLParser()
        parser.feed(html_content)
        
        invoices_html = []
        for table in parser.tables:
            if not table:
                continue
            headers = [h.strip() for h in table[0]]
            if 'N Fiscal' not in headers or 'Prorrog.' not in headers:
                continue
            
            n_fiscal_idx = headers.index('N Fiscal')
            prorrog_idx = headers.index('Prorrog.')
            
            for row in table[1:]:
                if len(row) <= max(n_fiscal_idx, prorrog_idx):
                    continue
                n_fiscal_raw = row[n_fiscal_idx].strip()
                prorrog_raw = row[prorrog_idx].strip()
                
                if not n_fiscal_raw or n_fiscal_raw.lower() == 'total' or 'total' in n_fiscal_raw.lower():
                    continue
                    
                # Trim leading 2 zeros (000017059 -> 0017059)
                n_fiscal_formatted = n_fiscal_raw[2:] if n_fiscal_raw.startswith('0000') else n_fiscal_raw
                
                invoices_html.append({
                    'raw_nf': n_fiscal_raw,
                    'nf': n_fiscal_formatted,
                    'prorrogacao': prorrog_raw
                })
                
        if not invoices_html:
            raise HTTPException(
                status_code=400,
                detail="Nenhum registro de prorrogação encontrado no arquivo HTML informado."
            )
            
        # 2. Parse CSV/Excel file
        csv_bytes = await csv_file.read()
        
        import io
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        import datetime

        # Load dynamic CSV/Excel
        try:
            df_csv = pd.read_excel(io.BytesIO(csv_bytes), header=None)
        except Exception:
            csv_text = csv_bytes.decode('utf-8', errors='ignore')
            sep = ';' if ';' in csv_text else ','
            df_csv = pd.read_csv(io.StringIO(csv_text), sep=sep, header=None)
            
        # Process and filter CSV data:
        # Col A (0): filter only 104
        # Col D (3): invoice number
        # Col E (4): parcela (must be 01)
        # Col P (15): due date (data de vencimento)
        csv_data_by_int = {}
        for idx, row in df_csv.iterrows():
            if len(row) < 16:
                continue
            val_a = str(row[0]).strip().split('.')[0]
            if val_a != '104':
                continue
            val_b = str(row[1]).strip().upper()
            if val_b != 'DP':
                continue
            val_d = str(row[3]).strip().split('.')[0]
            
            # Filtro da Coluna E (Parcela) - sempre 01
            val_e = str(row[4]).strip().split('.')[0]
            try:
                if int(val_e) != 1:
                    continue
            except ValueError:
                continue
                
            val_p = str(row[15]).strip()
            
            # Remove time component if present
            if ' ' in val_p:
                val_p = val_p.split()[0]
                
            try:
                csv_data_by_int[int(val_d)] = val_p
            except ValueError:
                continue
                
        # Helper to parse dates
        def parse_date(date_str):
            if not date_str:
                return None
            date_str = str(date_str).strip()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d/%m/%y', '%Y/%m/%d %H:%M:%S', '%d/%m/%Y %H:%M:%S'):
                try:
                    return datetime.datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue
            return date_str

        def format_date_to_br(date_str):
            if not date_str:
                return ""
            d = parse_date(date_str)
            if isinstance(d, datetime.date):
                return d.strftime('%d/%m/%Y')
            return str(date_str)

        # 3. Create Excel workbook with two sheets
        wb = openpyxl.Workbook()

        # Sheet 1: Prorrogações Atacadão
        ws1 = wb.active
        ws1.title = "Prorrogações Atacadão"
        ws1.views.sheetView[0].showGridLines = True

        ws1.cell(row=1, column=1, value="Nota Fiscal")
        ws1.cell(row=1, column=2, value="Data de Prorrogação")

        for i, inv in enumerate(invoices_html):
            row_num = i + 2
            ws1.cell(row=row_num, column=1, value=inv['nf'])
            ws1.cell(row=row_num, column=2, value=format_date_to_br(inv['prorrogacao']))

        # Sheet 2: Conciliação
        ws2 = wb.create_sheet(title="Conciliação")
        ws2.views.sheetView[0].showGridLines = True

        ws2.cell(row=1, column=1, value="Nota Fiscal")
        ws2.cell(row=1, column=2, value="Data HTML (Prorrogação)")
        ws2.cell(row=1, column=3, value="Data CSV (Vencimento)")
        ws2.cell(row=1, column=4, value="Status")

        # Fonts, alignments and colors
        font_header = Font(name='Calibri', size=11, bold=True)
        font_body = Font(name='Calibri', size=11, bold=False)
        align_left = Alignment(horizontal='left', vertical='center')
        align_center = Alignment(horizontal='center', vertical='center')
        
        fill_ok = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # soft green
        fill_div = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # soft orange

        # Format Sheet 1 Headers
        for col_idx in [1, 2]:
            cell = ws1.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left

        # Format Sheet 1 Body
        for r_idx in range(2, len(invoices_html) + 2):
            c1 = ws1.cell(row=r_idx, column=1)
            c1.font = font_body
            c1.number_format = '@'
            c1.alignment = align_left
            
            c2 = ws1.cell(row=r_idx, column=2)
            c2.font = font_body
            c2.alignment = align_center

        ws1.column_dimensions['A'].width = 16
        ws1.column_dimensions['B'].width = 24

        # Conciliate and populate Sheet 2
        for i, inv in enumerate(invoices_html):
            row_num = i + 2
            nf_str = inv['nf']
            date_html_str = inv['prorrogacao']
            
            try:
                nf_int = int(inv['raw_nf'])
            except ValueError:
                nf_int = None
                
            date_csv_str = ""
            status = "Não encontrado no CSV"
            status_fill = None
            
            if nf_int is not None and nf_int in csv_data_by_int:
                date_csv_str = csv_data_by_int[nf_int]
                d_html = parse_date(date_html_str)
                d_csv = parse_date(date_csv_str)
                
                # Check for equivalence or string match
                if d_html and d_csv:
                    if d_html == d_csv:
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                else:
                    if str(date_html_str).strip() == str(date_csv_str).strip():
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                        
            ws2.cell(row=row_num, column=1, value=nf_str)
            ws2.cell(row=row_num, column=2, value=format_date_to_br(date_html_str))
            ws2.cell(row=row_num, column=3, value=format_date_to_br(date_csv_str))
            
            status_cell = ws2.cell(row=row_num, column=4, value=status)
            if status_fill:
                status_cell.fill = status_fill

        # Format Sheet 2 Headers
        for col_idx in [1, 2, 3, 4]:
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left

        # Format Sheet 2 Body
        for r_idx in range(2, len(invoices_html) + 2):
            ws2.cell(row=r_idx, column=1).font = font_body
            ws2.cell(row=r_idx, column=1).number_format = '@'
            ws2.cell(row=r_idx, column=1).alignment = align_left
            
            ws2.cell(row=r_idx, column=2).font = font_body
            ws2.cell(row=r_idx, column=2).alignment = align_center
            
            ws2.cell(row=r_idx, column=3).font = font_body
            ws2.cell(row=r_idx, column=3).alignment = align_center
            
            ws2.cell(row=r_idx, column=4).font = font_body
            ws2.cell(row=r_idx, column=4).alignment = align_center

        ws2.column_dimensions['A'].width = 16
        ws2.column_dimensions['B'].width = 24
        ws2.column_dimensions['C'].width = 24
        ws2.column_dimensions['D'].width = 24

        # Save workbook to memory buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = html_file.filename.rsplit('.', 1)[0] + "_conciliado.xlsx"
        
        # Registrar no banco
        ImportacaoService(db).registrar_importacao(filename, "xlsx", "Prorrogação - Atacadão")
        
        headers_response = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_response
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/sendas/conciliar")
async def conciliar_sendas(
    sendas_file: UploadFile = File(...),
    acr_file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not sendas_file.filename or not acr_file.filename:
        raise HTTPException(status_code=400, detail="Arquivos inválidos")
        
    try:
        # Load Sendas file
        sendas_bytes = await sendas_file.read()
        import io
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        import datetime

        try:
            df_sendas = pd.read_excel(io.BytesIO(sendas_bytes), header=None)
        except Exception:
            csv_text = sendas_bytes.decode('utf-8', errors='ignore')
            sep = ';' if ';' in csv_text else ','
            df_sendas = pd.read_csv(io.StringIO(csv_text), sep=sep, header=None)

        if df_sendas.shape[1] < 13:
            raise HTTPException(
                status_code=400,
                detail=f"O arquivo Sendas precisa ter pelo menos 13 colunas (até a coluna M). Colunas encontradas: {df_sendas.shape[1]}."
            )

        # Parse Sendas data
        h_col_idx = 7 # Column H
        m_col_idx = 12 # Column M
        
        invoices_sendas = []
        for idx, row in df_sendas.iterrows():
            val_h_raw = row[h_col_idx]
            val_m_raw = row[m_col_idx]
            
            if pd.isna(val_h_raw) or pd.isna(val_m_raw):
                continue
                
            val_h = str(val_h_raw).strip()
            if val_h.endswith('.0'):
                val_h = val_h[:-2]
                
            # Filter starts with 1 or 2 and ends with 2
            if val_h.startswith(('1', '2')) and val_h.endswith('2'):
                # Format: remove spaces, remove final '2', pad to 7 characters
                cleaned_h = val_h.replace(' ', '')
                if cleaned_h.endswith('2'):
                    cleaned_h = cleaned_h[:-1]
                formatted_nf = cleaned_h.zfill(7)
                
                date_m = str(val_m_raw).strip()
                if ' ' in date_m:
                    date_m = date_m.split()[0]
                    
                invoices_sendas.append({
                    'raw_nf': val_h,
                    'nf': formatted_nf,
                    'vencimento': date_m
                })
                
        if not invoices_sendas:
            raise HTTPException(
                status_code=400,
                detail="Nenhum registro correspondente ao padrão Sendas foi encontrado na coluna H."
            )

        # Load ACR file
        acr_bytes = await acr_file.read()
        try:
            df_acr = pd.read_excel(io.BytesIO(acr_bytes), header=None)
        except Exception:
            csv_text = acr_bytes.decode('utf-8', errors='ignore')
            sep = ';' if ';' in csv_text else ','
            df_acr = pd.read_csv(io.StringIO(csv_text), sep=sep, header=None)

        if df_acr.shape[1] < 16:
            raise HTTPException(
                status_code=400,
                detail=f"O arquivo ACR precisa ter pelo menos 16 colunas (até a coluna P). Colunas encontradas: {df_acr.shape[1]}."
            )

        # Parse ACR data
        d_col_idx = 3 # Column D
        p_col_idx = 15 # Column P
        
        acr_data_by_int = {}
        for idx, row in df_acr.iterrows():
            val_d_raw = row[d_col_idx]
            val_p_raw = row[p_col_idx]
            
            if pd.isna(val_d_raw) or pd.isna(val_p_raw):
                continue
                
            val_d = str(val_d_raw).strip().split('.')[0]
            val_p = str(val_p_raw).strip()
            if ' ' in val_p:
                val_p = val_p.split()[0]
                
            try:
                acr_data_by_int[int(val_d)] = val_p
            except ValueError:
                continue

        # Helper to parse dates
        def parse_date(date_str):
            if not date_str:
                return None
            date_str = str(date_str).strip()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d/%m/%y', '%Y/%m/%d %H:%M:%S', '%d/%m/%Y %H:%M:%S'):
                try:
                    return datetime.datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue
            return date_str

        def format_date_to_br(date_str):
            if not date_str:
                return ""
            d = parse_date(date_str)
            if isinstance(d, datetime.date):
                return d.strftime('%d/%m/%Y')
            return str(date_str)

        # Create Workbook
        wb = openpyxl.Workbook()
        
        # Tab 1
        ws1 = wb.active
        ws1.title = "Prorrogações Sendas"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.cell(row=1, column=1, value="Nota Fiscal")
        ws1.cell(row=1, column=2, value="Data de Vencimento")
        
        for i, inv in enumerate(invoices_sendas):
            row_num = i + 2
            ws1.cell(row=row_num, column=1, value=inv['nf'])
            ws1.cell(row=row_num, column=2, value=format_date_to_br(inv['vencimento']))
            
        # Tab 2
        ws2 = wb.create_sheet(title="Conciliação")
        ws2.views.sheetView[0].showGridLines = True
        
        ws2.cell(row=1, column=1, value="Nota Fiscal")
        ws2.cell(row=1, column=2, value="Data Sendas (Vencimento)")
        ws2.cell(row=1, column=3, value="Data ACR (Vencimento)")
        ws2.cell(row=1, column=4, value="Status")

        # Fonts, alignments and colors
        font_header = Font(name='Calibri', size=11, bold=True)
        font_body = Font(name='Calibri', size=11, bold=False)
        align_left = Alignment(horizontal='left', vertical='center')
        align_center = Alignment(horizontal='center', vertical='center')
        
        fill_ok = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fill_div = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

        # Format Tab 1 Headers
        for col_idx in [1, 2]:
            cell = ws1.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left

        # Format Tab 1 Body
        for r_idx in range(2, len(invoices_sendas) + 2):
            c1 = ws1.cell(row=r_idx, column=1)
            c1.font = font_body
            c1.number_format = '@'
            c1.alignment = align_left
            
            c2 = ws1.cell(row=r_idx, column=2)
            c2.font = font_body
            c2.alignment = align_center

        ws1.column_dimensions['A'].width = 16
        ws1.column_dimensions['B'].width = 24

        # Conciliate
        for i, inv in enumerate(invoices_sendas):
            row_num = i + 2
            nf_str = inv['nf']
            date_sendas_str = inv['vencimento']
            
            try:
                raw_nf_cleaned = inv['raw_nf'].replace(' ', '')
                if raw_nf_cleaned.endswith('2'):
                    raw_nf_cleaned = raw_nf_cleaned[:-1]
                nf_int = int(raw_nf_cleaned)
            except ValueError:
                nf_int = None
                
            date_acr_str = ""
            status = "Não encontrado no ACR"
            status_fill = None
            
            if nf_int is not None and nf_int in acr_data_by_int:
                date_acr_str = acr_data_by_int[nf_int]
                d_sendas = parse_date(date_sendas_str)
                d_acr = parse_date(date_acr_str)
                
                if d_sendas and d_acr:
                    if d_sendas == d_acr:
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                else:
                    if str(date_sendas_str).strip() == str(date_acr_str).strip():
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                        
            ws2.cell(row=row_num, column=1, value=nf_str)
            ws2.cell(row=row_num, column=2, value=format_date_to_br(date_sendas_str))
            ws2.cell(row=row_num, column=3, value=format_date_to_br(date_acr_str))
            
            status_cell = ws2.cell(row=row_num, column=4, value=status)
            if status_fill:
                status_cell.fill = status_fill

        # Format Tab 2 Headers
        for col_idx in [1, 2, 3, 4]:
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left

        # Format Tab 2 Body
        for r_idx in range(2, len(invoices_sendas) + 2):
            ws2.cell(row=r_idx, column=1).font = font_body
            ws2.cell(row=r_idx, column=1).number_format = '@'
            ws2.cell(row=r_idx, column=1).alignment = align_left
            
            ws2.cell(row=r_idx, column=2).font = font_body
            ws2.cell(row=r_idx, column=2).alignment = align_center
            
            ws2.cell(row=r_idx, column=3).font = font_body
            ws2.cell(row=r_idx, column=3).alignment = align_center
            
            ws2.cell(row=r_idx, column=4).font = font_body
            ws2.cell(row=r_idx, column=4).alignment = align_center

        ws2.column_dimensions['A'].width = 16
        ws2.column_dimensions['B'].width = 24
        ws2.column_dimensions['C'].width = 24
        ws2.column_dimensions['D'].width = 24

        # Save Workbook
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = sendas_file.filename.rsplit('.', 1)[0] + "_conciliado.xlsx"
        
        # Registrar no banco
        ImportacaoService(db).registrar_importacao(filename, "xlsx", "Prorrogação - Sendas")
        
        headers_response = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_response
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/martminas/conciliar")
async def conciliar_martminas(
    martminas_file: UploadFile = File(...),
    acr_file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not martminas_file.filename or not acr_file.filename:
        raise HTTPException(status_code=400, detail="Arquivos inválidos")
        
    try:
        # Load Mart Minas file
        martminas_bytes = await martminas_file.read()
        import io
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        import datetime

        try:
            df_mm = pd.read_excel(io.BytesIO(martminas_bytes), header=None)
        except Exception:
            csv_text = martminas_bytes.decode('utf-8', errors='ignore')
            sep = ';' if ';' in csv_text else ','
            df_mm = pd.read_csv(io.StringIO(csv_text), sep=sep, header=None)

        if df_mm.shape[1] < 5:
            raise HTTPException(
                status_code=400,
                detail=f"O arquivo Mart Minas precisa ter pelo menos 5 colunas (até a coluna E). Colunas encontradas: {df_mm.shape[1]}."
            )

        # Parse Mart Minas data
        # Col E (4): nro documento
        # Col D (3): vencimento
        e_col_idx = 4
        d_col_idx = 3
        
        invoices_mm = []
        for idx, row in df_mm.iterrows():
            val_e_raw = row[e_col_idx]
            val_d_raw = row[d_col_idx]
            
            if pd.isna(val_e_raw) or pd.isna(val_d_raw):
                continue
                
            val_e = str(val_e_raw).strip()
            if val_e.endswith('.0'):
                val_e = val_e[:-2]
                
            # Skip header rows or non-numeric document numbers
            try:
                val_e_clean = val_e.replace(' ', '')
                if not val_e_clean:
                    continue
                int(val_e_clean)
            except ValueError:
                continue
                
            # Formatting: zfill to 7 characters
            formatted_nf = val_e.zfill(7)
            
            date_d = str(val_d_raw).strip()
            if ' ' in date_d:
                date_d = date_d.split()[0]
                
            invoices_mm.append({
                'raw_nf': val_e,
                'nf': formatted_nf,
                'vencimento': date_d
            })
            
        if not invoices_mm:
            raise HTTPException(
                status_code=400,
                detail="Nenhum registro correspondente ao padrão Mart Minas foi encontrado na coluna E."
            )

        # Load ACR file
        acr_bytes = await acr_file.read()
        try:
            df_acr = pd.read_excel(io.BytesIO(acr_bytes), header=None)
        except Exception:
            csv_text = acr_bytes.decode('utf-8', errors='ignore')
            sep = ';' if ';' in csv_text else ','
            df_acr = pd.read_csv(io.StringIO(csv_text), sep=sep, header=None)

        if df_acr.shape[1] < 16:
            raise HTTPException(
                status_code=400,
                detail=f"O arquivo ACR precisa ter pelo menos 16 colunas (até a coluna P). Colunas encontradas: {df_acr.shape[1]}."
            )

        # Parse ACR data
        acr_d_col_idx = 3 # Column D
        acr_p_col_idx = 15 # Column P
        
        acr_data_by_int = {}
        for idx, row in df_acr.iterrows():
            val_d_raw = row[acr_d_col_idx]
            val_p_raw = row[acr_p_col_idx]
            
            if pd.isna(val_d_raw) or pd.isna(val_p_raw):
                continue
                
            val_d = str(val_d_raw).strip().split('.')[0]
            val_p = str(val_p_raw).strip()
            if ' ' in val_p:
                val_p = val_p.split()[0]
                
            try:
                acr_data_by_int[int(val_d)] = val_p
            except ValueError:
                continue

        # Helper to parse dates
        def parse_date(date_str):
            if not date_str:
                return None
            date_str = str(date_str).strip()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d/%m/%y', '%Y/%m/%d %H:%M:%S', '%d/%m/%Y %H:%M:%S'):
                try:
                    return datetime.datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue
            return date_str

        def format_date_to_br(date_str):
            if not date_str:
                return ""
            d = parse_date(date_str)
            if isinstance(d, datetime.date):
                return d.strftime('%d/%m/%Y')
            return str(date_str)

        # Create Workbook
        wb = openpyxl.Workbook()
        
        # Tab 1
        ws1 = wb.active
        ws1.title = "Prorrogações Mart Minas"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.cell(row=1, column=1, value="Nota Fiscal")
        ws1.cell(row=1, column=2, value="Vencimento")
        
        for i, inv in enumerate(invoices_mm):
            row_num = i + 2
            ws1.cell(row=row_num, column=1, value=inv['nf'])
            ws1.cell(row=row_num, column=2, value=format_date_to_br(inv['vencimento']))
            
        # Tab 2
        ws2 = wb.create_sheet(title="Conciliação")
        ws2.views.sheetView[0].showGridLines = True
        
        ws2.cell(row=1, column=1, value="Nota Fiscal")
        ws2.cell(row=1, column=2, value="Vencimento Mart Minas")
        ws2.cell(row=1, column=3, value="Vencimento ACR")
        ws2.cell(row=1, column=4, value="Status")

        # Fonts, alignments and colors
        font_header = Font(name='Calibri', size=11, bold=True)
        font_body = Font(name='Calibri', size=11, bold=False)
        align_left = Alignment(horizontal='left', vertical='center')
        align_center = Alignment(horizontal='center', vertical='center')
        
        fill_ok = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fill_div = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

        # Format Tab 1 Headers
        for col_idx in [1, 2]:
            cell = ws1.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left

        # Format Tab 1 Body
        for r_idx in range(2, len(invoices_mm) + 2):
            c1 = ws1.cell(row=r_idx, column=1)
            c1.font = font_body
            c1.number_format = '@'
            c1.alignment = align_left
            
            c2 = ws1.cell(row=r_idx, column=2)
            c2.font = font_body
            c2.alignment = align_center

        ws1.column_dimensions['A'].width = 16
        ws1.column_dimensions['B'].width = 24

        # Conciliate
        for i, inv in enumerate(invoices_mm):
            row_num = i + 2
            nf_str = inv['nf']
            date_mm_str = inv['vencimento']
            
            try:
                raw_nf_cleaned = inv['raw_nf'].replace(' ', '')
                nf_int = int(raw_nf_cleaned)
            except ValueError:
                nf_int = None
                
            date_acr_str = ""
            status = "Não encontrado no ACR"
            status_fill = None
            
            if nf_int is not None and nf_int in acr_data_by_int:
                date_acr_str = acr_data_by_int[nf_int]
                d_mm = parse_date(date_mm_str)
                d_acr = parse_date(date_acr_str)
                
                if d_mm and d_acr:
                    if d_mm == d_acr:
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                else:
                    if str(date_mm_str).strip() == str(date_acr_str).strip():
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                        
            ws2.cell(row=row_num, column=1, value=nf_str)
            ws2.cell(row=row_num, column=2, value=format_date_to_br(date_mm_str))
            ws2.cell(row=row_num, column=3, value=format_date_to_br(date_acr_str))
            
            status_cell = ws2.cell(row=row_num, column=4, value=status)
            if status_fill:
                status_cell.fill = status_fill

        # Format Tab 2 Headers
        for col_idx in [1, 2, 3, 4]:
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left

        # Format Tab 2 Body
        for r_idx in range(2, len(invoices_mm) + 2):
            ws2.cell(row=r_idx, column=1).font = font_body
            ws2.cell(row=r_idx, column=1).number_format = '@'
            ws2.cell(row=r_idx, column=1).alignment = align_left
            
            ws2.cell(row=r_idx, column=2).font = font_body
            ws2.cell(row=r_idx, column=2).alignment = align_center
            
            ws2.cell(row=r_idx, column=3).font = font_body
            ws2.cell(row=r_idx, column=3).alignment = align_center
            
            ws2.cell(row=r_idx, column=4).font = font_body
            ws2.cell(row=r_idx, column=4).alignment = align_center

        ws2.column_dimensions['A'].width = 16
        ws2.column_dimensions['B'].width = 24
        ws2.column_dimensions['C'].width = 24
        ws2.column_dimensions['D'].width = 24

        # Save Workbook
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = martminas_file.filename.rsplit('.', 1)[0] + "_conciliado.xlsx"
        
        # Registrar no banco
        ImportacaoService(db).registrar_importacao(filename, "xlsx", "Prorrogação - Mart Minas")
        
        headers_response = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_response
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/savegnago/conciliar")
async def conciliar_savegnago(
    savegnago_file: UploadFile = File(...),
    acr_file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not savegnago_file.filename or not acr_file.filename:
        raise HTTPException(status_code=400, detail="Arquivos inválidos")
        
    try:
        # Load Savegnago file
        savegnago_bytes = await savegnago_file.read()
        import io
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        import datetime
        import re

        try:
            df_sav = pd.read_excel(io.BytesIO(savegnago_bytes), header=None)
        except Exception:
            csv_text = savegnago_bytes.decode('utf-8', errors='ignore')
            sep = ';' if ';' in csv_text else ','
            df_sav = pd.read_csv(io.StringIO(csv_text), sep=sep, header=None)

        if df_sav.shape[1] < 9:
            raise HTTPException(
                status_code=400,
                detail=f"O arquivo Savegnago precisa ter pelo menos 9 colunas (até a coluna I). Colunas encontradas: {df_sav.shape[1]}."
            )

        # Helper to parse dates
        def parse_date(date_str):
            if not date_str:
                return None
            if isinstance(date_str, datetime.datetime):
                return date_str.date()
            if isinstance(date_str, datetime.date):
                return date_str
            date_str = str(date_str).strip()
            if hasattr(pd, 'Timestamp') and isinstance(date_str, pd.Timestamp):
                return date_str.date()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d/%m/%y', '%Y/%m/%d %H:%M:%S', '%d/%m/%Y %H:%M:%S'):
                try:
                    return datetime.datetime.strptime(date_str.split()[0] if ' ' in date_str else date_str, fmt).date()
                except ValueError:
                    continue
            return None

        # Helper to adjust Savegnago date
        def adjust_savegnago_date(d_val):
            d = parse_date(d_val)
            if not isinstance(d, datetime.date):
                return d_val
            day = d.day
            month = d.month
            year = d.year
            
            if day == 31 or (1 <= day <= 9):
                new_day = 10
                if day == 31:
                    if month == 12:
                        month = 1
                        year += 1
                    else:
                        month += 1
            elif 11 <= day <= 19:
                new_day = 20
            elif 21 <= day <= 29:
                new_day = 30
            else:
                new_day = day
                
            try:
                adjusted = datetime.date(year, month, new_day)
                return adjusted
            except ValueError:
                if month == 2:
                    is_leap = (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0))
                    last_day = 29 if is_leap else 28
                    if new_day > last_day:
                        return datetime.date(year, month, last_day)
                return d

        def format_date_to_br(date_val):
            if not date_val:
                return ""
            if isinstance(date_val, datetime.date):
                return date_val.strftime('%d/%m/%Y')
            d = parse_date(date_val)
            if isinstance(d, datetime.date):
                return d.strftime('%d/%m/%Y')
            return str(date_val)

        # Parse Savegnago data
        d_col_idx = 3
        i_col_idx = 8
        
        invoices_sav = []
        for idx, row in df_sav.iterrows():
            val_d_raw = row[d_col_idx]
            val_i_raw = row[i_col_idx]
            
            if pd.isna(val_d_raw) or pd.isna(val_i_raw):
                continue
                
            val_d_str = str(val_d_raw).strip()
            if val_d_str.endswith('.0'):
                val_d_str = val_d_str[:-2]
                
            segment = val_d_str.split('-')[0]
            clean_segment = re.sub(r'^[a-zA-Z]+', '', segment)
            
            try:
                int(clean_segment)
            except ValueError:
                continue
                
            formatted_nf = clean_segment.zfill(7)
            
            parcela = ""
            qp_match = re.search(r'QP(\d+)', val_d_str)
            if qp_match:
                parcela = str(qp_match.group(1)).zfill(2)
            
            date_raw = str(val_i_raw).strip()
            if ' ' in date_raw:
                date_raw = date_raw.split()[0]
                
            adjusted_date = adjust_savegnago_date(date_raw)
            
            invoices_sav.append({
                'raw_nf': clean_segment,
                'nf': formatted_nf,
                'parcela': parcela,
                'vencimento': adjusted_date
            })
            
        if not invoices_sav:
            raise HTTPException(
                status_code=400,
                detail="Nenhum registro correspondente ao padrão Savegnago foi encontrado na coluna D."
            )

        # Load ACR file
        acr_bytes = await acr_file.read()
        try:
            df_acr = pd.read_excel(io.BytesIO(acr_bytes), header=None)
        except Exception:
            csv_text = acr_bytes.decode('utf-8', errors='ignore')
            sep = ';' if ';' in csv_text else ','
            df_acr = pd.read_csv(io.StringIO(csv_text), sep=sep, header=None)

        if df_acr.shape[1] < 16:
            raise HTTPException(
                status_code=400,
                detail=f"O arquivo ACR precisa ter pelo menos 16 colunas (até a coluna P). Colunas encontradas: {df_acr.shape[1]}."
            )

        # Parse ACR data
        acr_d_col_idx = 3 # Column D
        acr_e_col_idx = 4 # Column E (Parcela)
        acr_p_col_idx = 15 # Column P
        
        acr_data_by_key = {}
        for idx, row in df_acr.iterrows():
            if len(row) < 16:
                continue
                
            val_d_raw = row[acr_d_col_idx]
            val_e_raw = row[acr_e_col_idx]
            val_p_raw = row[acr_p_col_idx]
            
            if pd.isna(val_d_raw) or pd.isna(val_p_raw):
                continue
                
            val_d_str = str(val_d_raw).strip().split('.')[0]
            val_e_str = str(val_e_raw).strip().split('.')[0]
            
            val_e = val_e_str.zfill(2) if val_e_str.isdigit() else val_e_str
            
            val_p = str(val_p_raw).strip()
            if ' ' in val_p:
                val_p = val_p.split()[0]
                
            try:
                nf_int = int(val_d_str)
                acr_data_by_key[(nf_int, val_e)] = parse_date(val_p) or val_p
            except ValueError:
                continue

        # Create Workbook
        wb = openpyxl.Workbook()
        
        # Tab 1
        ws1 = wb.active
        ws1.title = "Prorrogações Savegnago"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.cell(row=1, column=1, value="Nota Fiscal")
        ws1.cell(row=1, column=2, value="Parcela")
        ws1.cell(row=1, column=3, value="Vencimento")
        
        for i, inv in enumerate(invoices_sav):
            row_num = i + 2
            ws1.cell(row=row_num, column=1, value=inv['nf'])
            ws1.cell(row=row_num, column=2, value=inv['parcela'])
            ws1.cell(row=row_num, column=3, value=format_date_to_br(inv['vencimento']))
            
        # Tab 2
        ws2 = wb.create_sheet(title="Conciliação")
        ws2.views.sheetView[0].showGridLines = True
        
        ws2.cell(row=1, column=1, value="Nota Fiscal")
        ws2.cell(row=1, column=2, value="Parcela")
        ws2.cell(row=1, column=3, value="Vencimento Savegnago")
        ws2.cell(row=1, column=4, value="Vencimento ACR")
        ws2.cell(row=1, column=5, value="Status")

        # Fonts, alignments and colors
        font_header = Font(name='Calibri', size=11, bold=True)
        font_body = Font(name='Calibri', size=11, bold=False)
        align_left = Alignment(horizontal='left', vertical='center')
        align_center = Alignment(horizontal='center', vertical='center')
        
        fill_ok = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fill_div = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

        # Format Tab 1 Headers
        for col_idx in [1, 2, 3]:
            cell = ws1.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left

        # Format Tab 1 Body
        for r_idx in range(2, len(invoices_sav) + 2):
            c1 = ws1.cell(row=r_idx, column=1)
            c1.font = font_body
            c1.number_format = '@'
            c1.alignment = align_left
            
            c2 = ws1.cell(row=r_idx, column=2)
            c2.font = font_body
            c2.number_format = '@'
            c2.alignment = align_center
            
            c3 = ws1.cell(row=r_idx, column=3)
            c3.font = font_body
            c3.alignment = align_center

        ws1.column_dimensions['A'].width = 16
        ws1.column_dimensions['B'].width = 10
        ws1.column_dimensions['C'].width = 24

        # Conciliate
        for i, inv in enumerate(invoices_sav):
            row_num = i + 2
            nf_str = inv['nf']
            parcela_str = inv['parcela']
            date_sav_val = inv['vencimento']
            
            try:
                nf_int = int(inv['raw_nf'])
            except ValueError:
                nf_int = None
                
            date_acr_str = ""
            status = "Não encontrado no ACR"
            status_fill = None
            
            search_key = (nf_int, parcela_str)
            if nf_int is not None and search_key in acr_data_by_key:
                date_acr_str = acr_data_by_key[search_key]
                
                if isinstance(date_sav_val, datetime.date):
                    d_sav = date_sav_val
                else:
                    d_sav = parse_date(date_sav_val)
                    
                d_acr = parse_date(date_acr_str) if not isinstance(date_acr_str, datetime.date) else date_acr_str
                
                if d_sav and d_acr:
                    if d_sav == d_acr:
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                else:
                    if str(format_date_to_br(date_sav_val)).strip() == str(format_date_to_br(date_acr_str)).strip():
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                        
            ws2.cell(row=row_num, column=1, value=nf_str)
            ws2.cell(row=row_num, column=2, value=parcela_str)
            ws2.cell(row=row_num, column=3, value=format_date_to_br(date_sav_val))
            ws2.cell(row=row_num, column=4, value=format_date_to_br(date_acr_str))
            
            status_cell = ws2.cell(row=row_num, column=5, value=status)
            if status_fill:
                status_cell.fill = status_fill

        # Format Tab 2 Headers
        for col_idx in [1, 2, 3, 4, 5]:
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left

        # Format Tab 2 Body
        for r_idx in range(2, len(invoices_sav) + 2):
            ws2.cell(row=r_idx, column=1).font = font_body
            ws2.cell(row=r_idx, column=1).number_format = '@'
            ws2.cell(row=r_idx, column=1).alignment = align_left
            
            ws2.cell(row=r_idx, column=2).font = font_body
            ws2.cell(row=r_idx, column=2).number_format = '@'
            ws2.cell(row=r_idx, column=2).alignment = align_center
            
            ws2.cell(row=r_idx, column=3).font = font_body
            ws2.cell(row=r_idx, column=3).alignment = align_center
            
            ws2.cell(row=r_idx, column=4).font = font_body
            ws2.cell(row=r_idx, column=4).alignment = align_center
            
            ws2.cell(row=r_idx, column=5).font = font_body
            ws2.cell(row=r_idx, column=5).alignment = align_center

        ws2.column_dimensions['A'].width = 16
        ws2.column_dimensions['B'].width = 10
        ws2.column_dimensions['C'].width = 24
        ws2.column_dimensions['D'].width = 24
        ws2.column_dimensions['E'].width = 24

        # Save Workbook
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = savegnago_file.filename.rsplit('.', 1)[0] + "_conciliado.xlsx"
        
        from app.services.importacao_service import ImportacaoService
        ImportacaoService(db).registrar_importacao(filename, "xlsx", "Prorrogação - Savegnago")
        
        headers_response = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_response
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/mateus/conciliar")
async def conciliar_mateus(
    mateus_file: UploadFile = File(...),
    acr_file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not mateus_file.filename or not acr_file.filename:
        raise HTTPException(status_code=400, detail="Arquivos inválidos")
        
    try:
        import io
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        import datetime
        import math
        
        # 1. Parse Mateus File
        mateus_bytes = await mateus_file.read()
        try:
            df_mat = pd.read_excel(io.BytesIO(mateus_bytes), header=None)
        except Exception:
            raise HTTPException(status_code=400, detail="Arquivo Mateus deve ser Excel.")
            
        def parse_date(date_str):
            if not date_str:
                return None
            if isinstance(date_str, datetime.datetime):
                return date_str.date()
            if isinstance(date_str, datetime.date):
                return date_str
            date_str = str(date_str).strip()
            if hasattr(pd, 'Timestamp') and isinstance(date_str, pd.Timestamp):
                return date_str.date()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%y', '%Y/%m/%d %H:%M:%S', '%d/%m/%Y %H:%M:%S'):
                try:
                    return datetime.datetime.strptime(date_str.split()[0] if ' ' in date_str else date_str, fmt).date()
                except ValueError:
                    continue
            return None

        mateus_invoices = []
        for idx, row in df_mat.iterrows():
            if len(row) < 6:
                continue
            
            nf_raw = str(row[2]).strip() # Col C (index 2)
            date_raw = row[5] # Col F (index 5)
            
            if not nf_raw or nf_raw.lower() in ('nan', 'none', 'nota fiscal', 'nota', 'nota_fiscal'):
                continue
                
            nf_clean = nf_raw.split('/')[0].strip()
            
            if not nf_clean:
                continue
                
            if not any(char.isdigit() for char in nf_clean):
                continue
            
            try:
                nf_clean = str(int(float(nf_clean))).zfill(7)
            except ValueError:
                nf_clean = nf_clean.zfill(7)
            
            d_parsed = parse_date(date_raw)
            d_adjusted = d_parsed if d_parsed else date_raw
            
            mateus_invoices.append({
                'raw_nf': nf_raw,
                'nf': nf_clean,
                'prorrogacao': d_adjusted
            })
            
        if not mateus_invoices:
            raise HTTPException(status_code=400, detail="Nenhuma nota fiscal encontrada no arquivo Mateus.")

        # 2. Parse ACR File
        acr_bytes = await acr_file.read()
        try:
            df_acr = pd.read_excel(io.BytesIO(acr_bytes), header=None)
        except Exception:
            acr_text = acr_bytes.decode('utf-8', errors='ignore')
            sep = ';' if ';' in acr_text else ','
            df_acr = pd.read_csv(io.StringIO(acr_text), sep=sep, header=None)
            
        acr_data_by_int = {}
        for idx, row in df_acr.iterrows():
            if len(row) < 16:
                continue
                
            val_d_raw = str(row[3]).strip()
            if not val_d_raw or val_d_raw.lower() in ('nan', 'none'):
                continue
                
            val_d = val_d_raw.split('.')[0]
            val_p = str(row[15]).strip()
            if ' ' in val_p:
                val_p = val_p.split()[0]
                
            try:
                acr_data_by_int[int(val_d)] = parse_date(val_p) or val_p
            except ValueError:
                continue
                
        # 3. Create Excel
        def format_date_to_br(d):
            if isinstance(d, datetime.date):
                return d.strftime('%d/%m/%Y')
            return str(d) if d else ""

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Prorrogações Mateus"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.cell(row=1, column=1, value="Nota Fiscal")
        ws1.cell(row=1, column=2, value="Data de Prorrogação")
        
        for i, inv in enumerate(mateus_invoices):
            ws1.cell(row=i+2, column=1, value=inv['nf'])
            ws1.cell(row=i+2, column=2, value=format_date_to_br(inv['prorrogacao']))
            
        ws2 = wb.create_sheet(title="Conciliação")
        ws2.views.sheetView[0].showGridLines = True
        ws2.cell(row=1, column=1, value="Nota Fiscal")
        ws2.cell(row=1, column=2, value="Data Mateus (Prorrogação)")
        ws2.cell(row=1, column=3, value="Data ACR (Vencimento)")
        ws2.cell(row=1, column=4, value="Status")
        
        font_header = Font(name='Calibri', size=11, bold=True)
        font_body = Font(name='Calibri', size=11, bold=False)
        align_left = Alignment(horizontal='left', vertical='center')
        align_center = Alignment(horizontal='center', vertical='center')
        
        fill_ok = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fill_div = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        for col_idx in [1, 2]:
            cell = ws1.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left
            
        for r_idx in range(2, len(mateus_invoices) + 2):
            ws1.cell(row=r_idx, column=1).font = font_body
            ws1.cell(row=r_idx, column=1).number_format = '@'
            ws1.cell(row=r_idx, column=1).alignment = align_left
            ws1.cell(row=r_idx, column=2).font = font_body
            ws1.cell(row=r_idx, column=2).alignment = align_center
            
        ws1.column_dimensions['A'].width = 16
        ws1.column_dimensions['B'].width = 24
        
        for i, inv in enumerate(mateus_invoices):
            row_num = i + 2
            nf_str = inv['nf']
            d_mat = inv['prorrogacao']
            
            try:
                nf_int = int(nf_str)
            except ValueError:
                nf_int = None
                
            d_acr = None
            status = "Não encontrado no ACR"
            status_fill = None
            
            if nf_int is not None and nf_int in acr_data_by_int:
                d_acr = acr_data_by_int[nf_int]
                
                if isinstance(d_mat, datetime.date) and isinstance(d_acr, datetime.date):
                    if d_mat == d_acr:
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                else:
                    if str(format_date_to_br(d_mat)).strip() == str(format_date_to_br(d_acr)).strip():
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                        
            ws2.cell(row=row_num, column=1, value=nf_str)
            ws2.cell(row=row_num, column=2, value=format_date_to_br(d_mat))
            ws2.cell(row=row_num, column=3, value=format_date_to_br(d_acr))
            
            status_cell = ws2.cell(row=row_num, column=4, value=status)
            if status_fill:
                status_cell.fill = status_fill

        for col_idx in [1, 2, 3, 4]:
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left
            
        for r_idx in range(2, len(mateus_invoices) + 2):
            ws2.cell(row=r_idx, column=1).font = font_body
            ws2.cell(row=r_idx, column=1).number_format = '@'
            ws2.cell(row=r_idx, column=1).alignment = align_left
            ws2.cell(row=r_idx, column=2).font = font_body
            ws2.cell(row=r_idx, column=2).alignment = align_center
            ws2.cell(row=r_idx, column=3).font = font_body
            ws2.cell(row=r_idx, column=3).alignment = align_center
            ws2.cell(row=r_idx, column=4).font = font_body
            ws2.cell(row=r_idx, column=4).alignment = align_center
            
        ws2.column_dimensions['A'].width = 16
        ws2.column_dimensions['B'].width = 28
        ws2.column_dimensions['C'].width = 28
        ws2.column_dimensions['D'].width = 24
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = mateus_file.filename.rsplit('.', 1)[0] + "_conciliado.xlsx"
        
        from app.services.importacao_service import ImportacaoService
        ImportacaoService(db).registrar_importacao(filename, "xlsx", "Prorrogação - Mateus")
        
        headers_response = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_response
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/drogaraia/conciliar")
async def conciliar_drogaraia(
    drogaraia_file: UploadFile = File(...),
    acr_file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not drogaraia_file.filename or not acr_file.filename:
        raise HTTPException(status_code=400, detail="Arquivos inválidos")
        
    try:
        import io
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        import datetime
        import math
        
        # 1. Parse Droga Raia File
        drogaraia_bytes = await drogaraia_file.read()
        try:
            df_dr = pd.read_excel(io.BytesIO(drogaraia_bytes), header=None)
        except Exception:
            try:
                dr_text = drogaraia_bytes.decode('utf-8', errors='ignore')
                sep = ';' if ';' in dr_text else ','
                df_dr = pd.read_csv(io.StringIO(dr_text), sep=sep, header=None)
            except Exception:
                try:
                    dr_text = drogaraia_bytes.decode('utf-8', errors='ignore')
                    dfs = pd.read_html(io.StringIO(dr_text), header=None)
                    df_dr = dfs[0]
                except Exception:
                    file_head = drogaraia_bytes[:100].decode('utf-8', errors='ignore')
                    with open("c:/Users/tania.canedo/.gemini/antigravity-ide/brain/612f4152-c197-44cc-a836-c8cadee18fba/scratch/file_format.log", "w") as f:
                        f.write(file_head)
                    raise HTTPException(status_code=400, detail=f"Arquivo desconhecido. Início do arquivo: {file_head[:50]}")
            
        def parse_date(date_str):
            if not date_str:
                return None
            if isinstance(date_str, datetime.datetime):
                return date_str.date()
            if isinstance(date_str, datetime.date):
                return date_str
            date_str = str(date_str).strip()
            if hasattr(pd, 'Timestamp') and isinstance(date_str, pd.Timestamp):
                return date_str.date()
                
            # Tratamento para datas em português ("4 de set de 2026")
            import re
            pt_months = {
                'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
                'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
            }
            pt_match = re.match(r'^(\d{1,2})\s+de\s+([a-zA-Z]{3,4})\s+de\s+(\d{4})$', date_str.lower())
            if pt_match:
                day = int(pt_match.group(1))
                month_str = pt_match.group(2)[:3]
                year = int(pt_match.group(3))
                if month_str in pt_months:
                    return datetime.date(year, pt_months[month_str], day)
                    
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%y', '%Y/%m/%d %H:%M:%S', '%d/%m/%Y %H:%M:%S'):
                try:
                    return datetime.datetime.strptime(date_str.split()[0] if ' ' in date_str else date_str, fmt).date()
                except ValueError:
                    continue
            return None

        drogaraia_invoices = []
        for idx, row in df_dr.iterrows():
            if len(row) < 7:
                continue
            
            nf_raw = str(row[4]).strip() # Col E (index 4)
            date_raw = row[6] # Col G (index 6)
            
            if not nf_raw or nf_raw.lower() in ('nan', 'none', 'nota fiscal', 'nota', 'nota_fiscal'):
                continue
                
            nf_clean = nf_raw.split('/')[0].strip()
            
            if not nf_clean:
                continue
                
            if not any(char.isdigit() for char in nf_clean):
                continue
            
            try:
                nf_clean = str(int(float(nf_clean))).zfill(7)
            except ValueError:
                nf_clean = nf_clean.zfill(7)
            
            d_parsed = parse_date(date_raw)
            d_adjusted = d_parsed if d_parsed else date_raw
            
            drogaraia_invoices.append({
                'raw_nf': nf_raw,
                'nf': nf_clean,
                'prorrogacao': d_adjusted
            })
            
        if not drogaraia_invoices:
            with open("c:/Users/tania.canedo/.gemini/antigravity-ide/brain/612f4152-c197-44cc-a836-c8cadee18fba/scratch/error.log", "a") as f:
                f.write(f"Droga Raia invoices empty. len(df_dr.columns)={len(df_dr.columns)}, rows={len(df_dr)}\n")
            raise HTTPException(status_code=400, detail="Nenhuma nota fiscal encontrada no arquivo Droga Raia.")

        # 2. Parse ACR File
        acr_bytes = await acr_file.read()
        try:
            df_acr = pd.read_excel(io.BytesIO(acr_bytes), header=None)
        except Exception:
            acr_text = acr_bytes.decode('utf-8', errors='ignore')
            sep = ';' if ';' in acr_text else ','
            df_acr = pd.read_csv(io.StringIO(acr_text), sep=sep, header=None)
            
        acr_data_by_int = {}
        for idx, row in df_acr.iterrows():
            if len(row) < 16:
                continue
                
            val_d_raw = str(row[3]).strip()
            if not val_d_raw or val_d_raw.lower() in ('nan', 'none'):
                continue
                
            val_d = val_d_raw.split('.')[0]
            val_p = str(row[15]).strip()
            if ' ' in val_p:
                val_p = val_p.split()[0]
                
            try:
                acr_data_by_int[int(val_d)] = parse_date(val_p) or val_p
            except ValueError:
                continue
                
        # 3. Create Excel
        def format_date_to_br(d):
            if isinstance(d, datetime.date):
                return d.strftime('%d/%m/%Y')
            return str(d) if d else ""

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Prorrogações Droga Raia"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.cell(row=1, column=1, value="Nota Fiscal")
        ws1.cell(row=1, column=2, value="Data de Prorrogação")
        
        for i, inv in enumerate(drogaraia_invoices):
            ws1.cell(row=i+2, column=1, value=inv['nf'])
            ws1.cell(row=i+2, column=2, value=format_date_to_br(inv['prorrogacao']))
            
        ws2 = wb.create_sheet(title="Conciliação")
        ws2.views.sheetView[0].showGridLines = True
        ws2.cell(row=1, column=1, value="Nota Fiscal")
        ws2.cell(row=1, column=2, value="Data Droga Raia (Prorrogação)")
        ws2.cell(row=1, column=3, value="Data ACR (Vencimento)")
        ws2.cell(row=1, column=4, value="Status")
        
        font_header = Font(name='Calibri', size=11, bold=True)
        font_body = Font(name='Calibri', size=11, bold=False)
        align_left = Alignment(horizontal='left', vertical='center')
        align_center = Alignment(horizontal='center', vertical='center')
        
        fill_ok = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fill_div = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        for col_idx in [1, 2]:
            cell = ws1.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left
            
        for r_idx in range(2, len(drogaraia_invoices) + 2):
            ws1.cell(row=r_idx, column=1).font = font_body
            ws1.cell(row=r_idx, column=1).number_format = '@'
            ws1.cell(row=r_idx, column=1).alignment = align_left
            ws1.cell(row=r_idx, column=2).font = font_body
            ws1.cell(row=r_idx, column=2).alignment = align_center
            
        ws1.column_dimensions['A'].width = 16
        ws1.column_dimensions['B'].width = 24
        
        for i, inv in enumerate(drogaraia_invoices):
            row_num = i + 2
            nf_str = inv['nf']
            d_dr = inv['prorrogacao']
            
            try:
                nf_int = int(nf_str)
            except ValueError:
                nf_int = None
                
            d_acr = None
            status = "Não encontrado no ACR"
            status_fill = None
            
            if nf_int is not None and nf_int in acr_data_by_int:
                d_acr = acr_data_by_int[nf_int]
                
                if isinstance(d_dr, datetime.date) and isinstance(d_acr, datetime.date):
                    if d_dr == d_acr:
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                else:
                    if str(format_date_to_br(d_dr)).strip() == str(format_date_to_br(d_acr)).strip():
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                        
            ws2.cell(row=row_num, column=1, value=nf_str)
            ws2.cell(row=row_num, column=2, value=format_date_to_br(d_dr))
            ws2.cell(row=row_num, column=3, value=format_date_to_br(d_acr))
            
            status_cell = ws2.cell(row=row_num, column=4, value=status)
            if status_fill:
                status_cell.fill = status_fill

        for col_idx in [1, 2, 3, 4]:
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left
            
        for r_idx in range(2, len(drogaraia_invoices) + 2):
            ws2.cell(row=r_idx, column=1).font = font_body
            ws2.cell(row=r_idx, column=1).number_format = '@'
            ws2.cell(row=r_idx, column=1).alignment = align_left
            ws2.cell(row=r_idx, column=2).font = font_body
            ws2.cell(row=r_idx, column=2).alignment = align_center
            ws2.cell(row=r_idx, column=3).font = font_body
            ws2.cell(row=r_idx, column=3).alignment = align_center
            ws2.cell(row=r_idx, column=4).font = font_body
            ws2.cell(row=r_idx, column=4).alignment = align_center
            
        ws2.column_dimensions['A'].width = 16
        ws2.column_dimensions['B'].width = 32
        ws2.column_dimensions['C'].width = 28
        ws2.column_dimensions['D'].width = 24
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = drogaraia_file.filename.rsplit('.', 1)[0] + "_conciliado.xlsx"
        
        from app.services.importacao_service import ImportacaoService
        ImportacaoService(db).registrar_importacao(filename, "xlsx", "Prorrogação - Droga Raia")
        
        headers_response = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_response
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/conciliacao-pagamentos/ler-apb")
async def ler_apb_planilha(file: UploadFile = File(...)):
    try:
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content), header=None)

        if df.shape[1] < 5:
            raise HTTPException(
                status_code=400,
                detail="A planilha APB deve conter pelo menos 5 colunas (colunas D e E)."
            )

        # Column D (index 3) is values
        df[3] = pd.to_numeric(df[3], errors='coerce').fillna(0.0)

        # Clean names and map invalid/ND placeholders to "Erro"
        def clean_and_map_name(val):
            if val is None or pd.isna(val):
                return "Erro"
            val_str = str(val).strip()
            val_upper = val_str.upper()
            # Ignore headers completely
            if val_upper in ["FORNECEDOR", "NOME", "FAVORECIDO", "BENEFICIÁRIO", "BENEFICIARIO", "NOME COMPLETO"]:
                return None
            # Map ND and empty items to "Erro"
            if val_upper in ["ND", "N/D", "NAN", "", "NONE", "NULL"]:
                return "Erro"
            return val_str

        result = []
        seen_valid = {}  # maps valid_name -> index in result_list

        for _, row in df.iterrows():
            name = clean_and_map_name(row[4])
            if name is None:
                continue

            value = float(row[3])

            if name == "Erro":
                result.append({
                    "nome": "Erro",
                    "valor": value
                })
            else:
                if name in seen_valid:
                    idx = seen_valid[name]
                    result[idx]["valor"] += value
                else:
                    seen_valid[name] = len(result)
                    result.append({
                        "nome": name,
                        "valor": value
                    })

        # Print list to Python console as requested
        print("\n=== LISTA DE PESSOAS CONSOLIDADAS (PLANILHA APB) ===")
        for idx, item in enumerate(result):
            print(f"{idx+1}. {item['nome']}: R$ {item['valor']:.2f}")
        print("====================================================\n")

        return {"sucesso": True, "dados": result}
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

from datetime import datetime
from typing import List

@router.post("/conciliacao-pagamentos/cruzamento")
async def conciliar_pagamentos_cruzamento(
    apb_file: UploadFile = File(...),
    banco_files: List[UploadFile] = File(...)
):
    try:
        print(f"[DEBUG CONCILIACAO] Iniciando processamento. APB: {apb_file.filename}, Arquivos de banco: {[f.filename for f in banco_files]}")
        # 1. Parse APB File
        content_apb = await apb_file.read()
        df_apb = pd.read_excel(io.BytesIO(content_apb))
        print(f"[DEBUG CONCILIACAO] Planilha APB carregada com {len(df_apb)} linhas e colunas: {list(df_apb.columns)}")
        
        def find_column_by_name(df, candidates):
            for col in df.columns:
                col_lower = str(col).lower().strip()
                if any(c in col_lower for c in candidates):
                    return col
            return None

        apb_list = []
        if not df_apb.empty:
            # 1. Tenta encontrar colunas por nome
            col_data = find_column_by_name(df_apb, ['data', 'vencimento', 'pagamento'])
            col_doc = find_column_by_name(df_apb, ['documento', 'titulo', 'numero', 'doc', 'nº'])
            col_forn = find_column_by_name(df_apb, ['fornecedor', 'favorecido', 'nome', 'beneficiario', 'cliente'])
            col_val = find_column_by_name(df_apb, ['valor', 'val', 'liquido', 'total', 'quant'])
            col_cnpj_cpf = find_column_by_name(df_apb, ['cnpj', 'cpf', 'identificacao'])
            col_conta = find_column_by_name(df_apb, ['conta', 'agencia', 'banco', 'pix', 'dados bancarios'])
            
            # 2. Se falhar, analisa o conteúdo real das células
            sample_size = min(10, len(df_apb))
            if sample_size > 0:
                column_types = {}
                for col in df_apb.columns:
                    non_null_vals = df_apb[col].dropna().head(sample_size).tolist()
                    if not non_null_vals:
                        continue
                    
                    is_numeric_count = 0
                    is_date_count = 0
                    is_cnpj_cpf_count = 0
                    avg_len = 0
                    
                    for val in non_null_vals:
                        val_str = str(val).strip()
                        avg_len += len(val_str)
                        
                        if isinstance(val, (datetime, pd.Timestamp)):
                            is_date_count += 1
                            continue
                        if re.match(r'^\d{2}/\d{2}/\d{2,4}$', val_str):
                            is_date_count += 1
                            continue
                            
                        digits = re.sub(r'\D', '', val_str)
                        if len(digits) in [11, 14] and digits.isdigit():
                            is_cnpj_cpf_count += 1
                            
                        cleaned_val = val_str.replace('R$', '').replace('.', '').replace(',', '.').replace('\xa0', '').strip()
                        try:
                            float(cleaned_val)
                            is_numeric_count += 1
                        except ValueError:
                            pass
                    
                    avg_len /= len(non_null_vals)
                    column_types[col] = {
                        "is_numeric_pct": is_numeric_count / len(non_null_vals),
                        "is_date_pct": is_date_count / len(non_null_vals),
                        "is_cnpj_cpf_pct": is_cnpj_cpf_count / len(non_null_vals),
                        "avg_len": avg_len,
                        "sample_vals": [str(x) for x in non_null_vals]
                    }
                
                # Mapeia colunas baseado na probabilidade
                if not col_val:
                    best_val_col = None
                    best_pct = 0.0
                    for col, info in column_types.items():
                        if info["is_date_pct"] > 0.5:
                            continue
                        if info["is_numeric_pct"] > best_pct:
                            best_pct = info["is_numeric_pct"]
                            best_val_col = col
                    if best_pct > 0.4:
                        col_val = best_val_col
                
                if not col_data:
                    best_date_col = None
                    best_pct = 0.0
                    for col, info in column_types.items():
                        if info["is_date_pct"] > best_pct:
                            best_pct = info["is_date_pct"]
                            best_date_col = col
                    if best_pct > 0.4:
                        col_data = best_date_col
                
                if not col_cnpj_cpf:
                    best_cnpj_col = None
                    best_pct = 0.0
                    for col, info in column_types.items():
                        if info["is_cnpj_cpf_pct"] > best_pct:
                            best_pct = info["is_cnpj_cpf_pct"]
                            best_cnpj_col = col
                    if best_pct > 0.4:
                        col_cnpj_cpf = best_cnpj_col
                
                if not col_forn:
                    best_forn_col = None
                    max_len = 0
                    for col, info in column_types.items():
                        if col in [col_val, col_data, col_cnpj_cpf, col_conta]:
                            continue
                        if info["avg_len"] <= 4 and any("R$" in val for val in info["sample_vals"]):
                            continue
                        if info["avg_len"] > max_len:
                            max_len = info["avg_len"]
                            best_forn_col = col
                    col_forn = best_forn_col

            # Definição final de fallbacks caso continue nulo
            col_val = col_val or (df_apb.columns[3] if len(df_apb.columns) > 3 else df_apb.columns[0])
            col_forn = col_forn or (df_apb.columns[2] if len(df_apb.columns) > 2 else df_apb.columns[0])
            col_data = col_data or (df_apb.columns[0] if len(df_apb.columns) > 0 else df_apb.columns[0])
            col_doc = col_doc or (df_apb.columns[1] if len(df_apb.columns) > 1 else df_apb.columns[0])
            
            print(f"[DEBUG CONCILIACAO] Colunas mapeadas - Valor: {col_val}, Fornecedor: {col_forn}, Data: {col_data}, Doc: {col_doc}")
            
            for idx, row in df_apb.iterrows():
                try:
                    val_raw = row[col_val]
                    if pd.isna(val_raw):
                        continue
                    
                    # Convert to float
                    val_str = str(val_raw).replace('R$', '').replace('.', '').replace(',', '.').replace('\xa0', '').strip()
                    val = float(val_str)
                    
                    date_raw = row[col_data] if col_data in df_apb.columns else None
                    date_str = ""
                    if pd.notna(date_raw):
                        if isinstance(date_raw, datetime):
                            date_str = date_raw.strftime("%d/%m/%Y")
                        else:
                            date_str = str(date_raw).strip()
                    
                    doc_raw = row[col_doc] if col_doc in df_apb.columns else ""
                    doc_str = str(doc_raw).strip() if pd.notna(doc_raw) else ""
                    if doc_str.endswith(".0"):
                        doc_str = doc_str[:-2]
                        
                    forn_raw = row[col_forn] if col_forn in df_apb.columns else ""
                    forn_str = str(forn_raw).strip() if pd.notna(forn_raw) else ""
                    
                    cnpj_cpf_str = ""
                    if col_cnpj_cpf and col_cnpj_cpf in df_apb.columns:
                        cnpj_cpf_raw = row[col_cnpj_cpf]
                        cnpj_cpf_str = str(cnpj_cpf_raw).strip() if pd.notna(cnpj_cpf_raw) else ""
                        if cnpj_cpf_str.endswith(".0"):
                            cnpj_cpf_str = cnpj_cpf_str[:-2]
                    
                    conta_str = ""
                    if col_conta and col_conta in df_apb.columns:
                        conta_raw = row[col_conta]
                        conta_str = str(conta_raw).strip() if pd.notna(conta_raw) else ""
                        if conta_str.endswith(".0"):
                            conta_str = conta_str[:-2]
                    
                    apb_list.append({
                        "data": date_str,
                        "documento": doc_str,
                        "fornecedor": forn_str,
                        "valor": val,
                        "cnpj_cpf": cnpj_cpf_str,
                        "dados_bancarios": conta_str,
                        "matched": False
                    })
                except Exception:
                    continue

        print(f"[DEBUG CONCILIACAO] Total de títulos carregados do APB: {len(apb_list)}")

        # 2. Parse Bank Files
        banco_list = []
        ia = IAService()
        
        for b_file in banco_files:
            b_filename = b_file.filename or "extrato"
            print(f"[DEBUG CONCILIACAO] Analisando arquivo de banco: {b_filename}")
            b_content = await b_file.read()
            
            if b_filename.lower().endswith(".pdf"):
                # Call Gemini for PDF extraction
                res_ia = await ia.analisar_banco_pdf(b_content, b_filename)
                extracted_trns = res_ia.get("transacoes", [])
                for tx in extracted_trns:
                    sit = tx.get("situacao", "")
                    if not sit:
                        desc_t = tx.get("descricao", "")
                        if "Confirmação de Pagamento" in desc_t or "PIX CEF MATRIZ" in desc_t:
                            sit = "BB-PIX"
                        else:
                            sit = "BB-LIB"
                    banco_list.append({
                        "data": tx.get("data", ""),
                        "descricao": tx.get("descricao", ""),
                        "valor": tx.get("valor", 0.0),
                        "documento": tx.get("documento", ""),
                        "situacao": sit
                    })
                print(f"[DEBUG CONCILIACAO] Extraídas {len(extracted_trns)} transações do PDF pelo Gemini.")
            elif b_filename.lower().endswith(".ofx"):
                # Parse OFX
                content_str = b_content.decode("utf-8", errors="ignore")
                transactions = []
                stmttrns = re.findall(r'<STMTTRN>(.*?)</STMTTRN>', content_str, re.DOTALL | re.IGNORECASE)
                if not stmttrns:
                    stmttrns = re.split(r'<STMTTRN>', content_str, flags=re.IGNORECASE)[1:]
                    
                for trn in stmttrns:
                    fitid = re.search(r'<FITID>(.*?)(?:\n|\r|<)', trn, re.IGNORECASE)
                    dtposted = re.search(r'<DTPOSTED>(.*?)(?:\n|\r|<)', trn, re.IGNORECASE)
                    trnamt = re.search(r'<TRNAMT>(.*?)(?:\n|\r|<)', trn, re.IGNORECASE)
                    memo = re.search(r'<MEMO>(.*?)(?:\n|\r|<)', trn, re.IGNORECASE)
                    name = re.search(r'<NAME>(.*?)(?:\n|\r|<)', trn, re.IGNORECASE)
                    
                    doc = fitid.group(1).strip() if fitid else ""
                    date_str = ""
                    if dtposted:
                        raw_dt = dtposted.group(1).strip()
                        if len(raw_dt) >= 8:
                            date_str = f"{raw_dt[6:8]}/{raw_dt[4:6]}/{raw_dt[0:4]}"
                    amt = 0.0
                    if trnamt:
                        try:
                            amt = abs(float(trnamt.group(1).strip()))
                        except ValueError:
                            pass
                    desc = memo.group(1).strip() if memo else (name.group(1).strip() if name else "Lançamento Bancário")
                    
                    sit = "BB-PIX" if ("Confirmação de Pagamento" in desc or "PIX CEF" in desc) else "BB-LIB"
                    transactions.append({
                        "data": date_str,
                        "descricao": desc,
                        "valor": amt,
                        "documento": doc,
                        "situacao": sit
                    })
                banco_list.extend(transactions)
            else:
                # Excel or CSV
                try:
                    if b_filename.lower().endswith(".csv"):
                        df_b = pd.read_csv(io.BytesIO(b_content))
                    else:
                        df_b = pd.read_excel(io.BytesIO(b_content))
                    
                    if not df_b.empty:
                        col_b_data = find_column(df_b, ['data', 'vencimento', 'pagamento']) or df_b.columns[0]
                        col_b_desc = find_column(df_b, ['descricao', 'historico', 'memo', 'detalhe', 'fornecedor', 'nome']) or (df_b.columns[1] if len(df_b.columns) > 1 else df_b.columns[0])
                        col_b_val = find_column(df_b, ['valor', 'val', 'lancamento', 'total', 'quant']) or (df_b.columns[2] if len(df_b.columns) > 2 else df_b.columns[0])
                        col_b_doc = find_column(df_b, ['documento', 'titulo', 'numero', 'doc'])
                        
                        for idx, row in df_b.iterrows():
                            try:
                                val_raw = row[col_b_val]
                                if pd.isna(val_raw):
                                    continue
                                val = abs(float(str(val_raw).replace('R$', '').replace('.', '').replace(',', '.').replace('\xa0', '').strip()))
                                
                                date_raw = row[col_b_data]
                                date_str = ""
                                if pd.notna(date_raw):
                                    if isinstance(date_raw, datetime):
                                        date_str = date_raw.strftime("%d/%m/%Y")
                                    else:
                                        date_str = str(date_raw).strip()
                                
                                doc_str = str(row[col_b_doc]).strip() if (col_b_doc and pd.notna(row[col_b_doc])) else ""
                                if doc_str.endswith(".0"):
                                    doc_str = doc_str[:-2]
                                desc_str = str(row[col_b_desc]).strip() if pd.notna(row[col_b_desc]) else "Lançamento Bancário"
                                
                                sit = "BB-PIX" if ("Confirmação de Pagamento" in desc_str or "PIX CEF" in desc_str) else "BB-LIB"
                                banco_list.append({
                                    "data": date_str,
                                    "descricao": desc_str,
                                    "valor": val,
                                    "documento": doc_str,
                                    "situacao": sit
                                })
                            except Exception:
                                continue
                except Exception:
                    continue

        print(f"[DEBUG CONCILIACAO] Total de transações bancárias carregadas: {len(banco_list)}")
        print("[DEBUG CONCILIACAO] Iniciando cruzamento de dados...")
        
        # 3. Cruzamento / Reconciliação
        matches = []
        for b_item in banco_list:
            matched_apb = None
            status = "nao_encontrado"
            
            # Regra Específica 1: PIX CEF MATRIZ -> FGTS FOLHA
            desc_b = b_item["descricao"].lower()
            if "pix cef matriz" in desc_b or "pagamento instantâneo-pix cef" in desc_b or "pagamento instantaneo-pix cef" in desc_b:
                for a_item in apb_list:
                    if not a_item["matched"] and abs(a_item["valor"] - b_item["valor"]) < 0.01:
                        forn_lower = a_item["fornecedor"].lower()
                        doc_lower = a_item["documento"].lower()
                        if "fgts" in forn_lower or "folha" in forn_lower or "fgts" in doc_lower or "folha" in doc_lower:
                            matched_apb = a_item
                            status = "conciliado"
                            b_item["situacao"] = "BB-PIX"
                            break
            
            # Match 1: documento + valor exato
            if not matched_apb and b_item["documento"]:
                for a_item in apb_list:
                    if not a_item["matched"] and a_item["documento"] == b_item["documento"] and abs(a_item["valor"] - b_item["valor"]) < 0.01:
                        matched_apb = a_item
                        status = "conciliado"
                        break
                        
            # Match 2: nome/CNPJ + valor exato + data aproximada (dentro de 5 dias)
            if not matched_apb:
                for a_item in apb_list:
                    if not a_item["matched"] and abs(a_item["valor"] - b_item["valor"]) < 0.01:
                        desc_lower = b_item["descricao"].lower()
                        forn_lower = a_item["fornecedor"].lower()
                        
                        name_match = (forn_lower in desc_lower or desc_lower in forn_lower or (a_item["cnpj_cpf"] and a_item["cnpj_cpf"] in desc_lower))
                        if name_match:
                            try:
                                b_date = datetime.strptime(b_item["data"], "%d/%m/%Y")
                                a_date = datetime.strptime(a_item["data"], "%d/%m/%Y")
                                if abs((b_date - a_date).days) <= 5:
                                    matched_apb = a_item
                                    status = "conciliado"
                                    break
                            except Exception:
                                matched_apb = a_item
                                status = "conciliado"
                                break
                                
            # Match 3: valor exato + data aproximada (fall-back geral)
            if not matched_apb:
                for a_item in apb_list:
                    if not a_item["matched"] and abs(a_item["valor"] - b_item["valor"]) < 0.01:
                        try:
                            b_date = datetime.strptime(b_item["data"], "%d/%m/%Y")
                            a_date = datetime.strptime(a_item["data"], "%d/%m/%Y")
                            if abs((b_date - a_date).days) <= 5:
                                matched_apb = a_item
                                status = "conciliado"
                                break
                        except Exception:
                            matched_apb = a_item
                            status = "conciliado"
                            break

            # Match 4 (Novo): Agrupamento de múltiplos títulos do APB que somam o valor total debitado/creditado no banco
            if not matched_apb:
                desc_lower = b_item["descricao"].lower()
                candidatos_apb = []
                for a_item in apb_list:
                    if not a_item["matched"]:
                        forn_lower = a_item["fornecedor"].lower()
                        if forn_lower in desc_lower or desc_lower in forn_lower or (a_item["cnpj_cpf"] and a_item["cnpj_cpf"] in desc_lower):
                            candidatos_apb.append(a_item)
                
                if candidatos_apb:
                    soma_valores = sum(c["valor"] for c in candidatos_apb)
                    if abs(soma_valores - b_item["valor"]) < 0.01:
                        for c in candidatos_apb:
                            c["matched"] = True
                        
                        matches.append({
                            "banco_data": b_item["data"],
                            "banco_descricao": b_item["descricao"],
                            "banco_documento": b_item["documento"],
                            "banco_valor": b_item["valor"],
                            "apb_data": ", ".join(list(set([c["data"] for c in candidatos_apb if c["data"]]))),
                            "apb_documento": ", ".join([c["documento"] for c in candidatos_apb if c["documento"]]),
                            "apb_fornecedor": candidatos_apb[0]["fornecedor"],
                            "apb_valor": soma_valores,
                            "cnpj_cpf": candidatos_apb[0]["cnpj_cpf"],
                            "dados_bancarios": candidatos_apb[0]["dados_bancarios"],
                            "status": "conciliado",
                            "situacao": b_item.get("situacao", "BB-LIB")
                        })
                        continue

            if matched_apb:
                matched_apb["matched"] = True
                matches.append({
                    "banco_data": b_item["data"],
                    "banco_descricao": b_item["descricao"],
                    "banco_documento": b_item["documento"],
                    "banco_valor": b_item["valor"],
                    "apb_data": matched_apb["data"],
                    "apb_documento": matched_apb["documento"],
                    "apb_fornecedor": matched_apb["fornecedor"],
                    "apb_valor": matched_apb["valor"],
                    "cnpj_cpf": matched_apb["cnpj_cpf"],
                    "dados_bancarios": matched_apb["dados_bancarios"],
                    "status": status,
                    "situacao": b_item.get("situacao", "BB-LIB")
                })
            else:
                matches.append({
                    "banco_data": b_item["data"],
                    "banco_descricao": b_item["descricao"],
                    "banco_documento": b_item["documento"],
                    "banco_valor": b_item["valor"],
                    "apb_data": "",
                    "apb_documento": "",
                    "apb_fornecedor": "",
                    "apb_valor": 0.0,
                    "cnpj_cpf": "",
                    "dados_bancarios": "",
                    "status": "nao_encontrado",
                    "situacao": b_item.get("situacao", "BB-LIB")
                })

        for a_item in apb_list:
            if not a_item["matched"]:
                matches.append({
                    "banco_data": "",
                    "banco_descricao": "",
                    "banco_documento": "",
                    "banco_valor": 0.0,
                    "apb_data": a_item["data"],
                    "apb_documento": a_item["documento"],
                    "apb_fornecedor": a_item["fornecedor"],
                    "apb_valor": a_item["valor"],
                    "cnpj_cpf": a_item["cnpj_cpf"],
                    "dados_bancarios": a_item["dados_bancarios"],
                    "status": "nao_encontrado",
                    "situacao": ""
                })

        print(f"[DEBUG CONCILIACAO] Cruzamento concluído com sucesso. Total de registros para conferência: {len(matches)}")
        return {"sucesso": True, "conferencia": matches}
    except Exception as e:
        import traceback
        print("[ERROR CONCILIACAO] Ocorreu uma exceção durante o cruzamento:")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

class ExportarConciliacaoRequest(BaseModel):
    dados: List[dict]
    fileName: Optional[str] = "conciliacao_pagamentos.xlsx"

@router.post("/conciliacao-pagamentos/exportar")
def exportar_conciliacao(
    req: ExportarConciliacaoRequest,
    db: Session = Depends(get_db)
):
    try:
        dados = req.dados
        filename = req.fileName or "conciliacao_pagamentos.xlsx"

        # Calculate stats
        matched = sum(1 for d in dados if d.get("status") == "conciliado")
        unmatched = sum(1 for d in dados if d.get("status") == "nao_encontrado")
        divergent = sum(1 for d in dados if d.get("status") == "divergente")
        status_val = "warning" if divergent > 0 else "success"
        user_val = "Ana Paula (Financeiro)"

        # Save Importacao record in database
        tipo_str = f"CONCILIACAO_BANCARIA|{matched}|{unmatched}|{divergent}|{status_val}|{user_val}"
        ImportacaoService(db).registrar_importacao(filename, "xlsx", tipo_str)

        df = pd.DataFrame(dados)
        df_rename = df.rename(columns={
            "banco_data": "Data Banco",
            "banco_descricao": "Descrição Banco",
            "banco_documento": "Doc Banco",
            "banco_valor": "Valor Banco",
            "apb_data": "Data APB",
            "apb_documento": "Doc APB",
            "apb_fornecedor": "Fornecedor APB",
            "apb_valor": "Valor APB",
            "cnpj_cpf": "CNPJ/CPF APB",
            "dados_bancarios": "Dados Bancários APB",
            "status": "Status Conciliação",
            "situacao": "Situação Banco"
        })
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_rename.to_excel(writer, index=False, sheet_name='Conciliação')
            
        output.seek(0)
        
        headers_response = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_response
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/plano-saude/sorriso/analisar")
async def analisar_plano_saude_sorriso(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo inválido")
        
    try:
        content = await file.read()
        
        # Obter todos os colaboradores do banco de dados para passar como candidatos ao Gemini
        colab_repo = ColaboradorRepository(db)
        # Buscar colaboradores do banco de dados
        colabs_db, _ = colab_repo.get_all(limit=5000)
        nomes_colaboradores = [c.nome for c in colabs_db]
        
        ia = IAService()
        res_ia = await ia.analisar_plano_saude_sorriso(
            file_content=content,
            file_name=file.filename,
            colaboradores=nomes_colaboradores
        )
        
        titulares_extraidos = res_ia.get("titulares", [])
        
        import difflib
        from app.repositories.colaborador_alias_repository import ColaboradorAliasRepository
        alias_repo = ColaboradorAliasRepository(db)
        
        # Injetar o Centro de Custo correspondente do banco para cada titular
        for t in titulares_extraidos:
            nome_pdf = t.get("nome_pdf", "")
            
            # 1. Verifica na tabela de Alias (aprendizado)
            alias_record = alias_repo.get_by_nome_divergente(nome_pdf)
            if alias_record and alias_record.colaborador:
                nome_db = alias_record.colaborador.nome
            else:
                # 2. Tentar achar o nome mais parecido no banco para evitar falsos negativos e falhas da IA
                nome_db = nome_pdf
                closest = difflib.get_close_matches(nome_pdf, nomes_colaboradores, n=1, cutoff=0.8)
                if closest:
                    nome_db = closest[0]
                
            t["nome_db"] = nome_db
            
            colab = colab_repo.get_by_nome(nome_db)
            if not colab:
                colab = colab_repo.get_by_nome(nome_pdf)
                
            if colab and colab.centro_custo:
                t["centro_custo"] = str(colab.centro_custo.codigo)
            else:
                t["centro_custo"] = "N/D"
        
        # Validations in python
        # 1. Confirmar que cada titular aparece apenas uma vez
        nomes_titulares = [t["nome_pdf"].strip().upper() for t in titulares_extraidos]
        titulares_unicos = set(nomes_titulares)
        
        # 2. Verificar que nenhum dependente foi listado como titular
        nomes_dependentes = []
        for t in titulares_extraidos:
            for d in t.get("dependentes", []):
                nomes_dependentes.append(d["nome"].strip().upper())
                
        intersection = titulares_unicos.intersection(set(nomes_dependentes))
        
        # 3. Confirmar que a soma dos valores individuais é igual ao TOTAL GERAL
        soma_individual = 0.0
        soma_grupo = 0.0
        for t in titulares_extraidos:
            val_tit = t["valor_titular"]
            val_deps = sum(d["valor"] for d in t.get("dependentes", []))
            soma_individual += val_tit + val_deps
            soma_grupo += t["valor_total"]
            
        # Validations object for frontend badges/alerts
        validacoes = {
            "apenas_titulares_na_tabela": True,
            "sem_titulares_duplicados": len(nomes_titulares) == len(titulares_unicos),
            "sem_dependentes_como_titulares": len(intersection) == 0,
            "soma_individual_bate_com_total_geral": round(soma_individual, 2) == round(soma_grupo, 2),
            "titulares_count": len(titulares_extraidos),
            "dependentes_count": len(nomes_dependentes),
            "total_count": len(titulares_extraidos) + len(nomes_dependentes)
        }
        
        validacoes_sucesso = all([
            validacoes["sem_titulares_duplicados"],
            validacoes["sem_dependentes_como_titulares"],
            validacoes["soma_individual_bate_com_total_geral"]
        ])
        
        return {
            "sucesso": True,
            "dados": titulares_extraidos,
            "validacoes": validacoes,
            "validacoes_sucesso": validacoes_sucesso,
            "total_geral": round(soma_grupo, 2)
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


class DependentConfirmar(BaseModel):
    nome: str
    valor: float

class TitularConfirmar(BaseModel):
    nome_pdf: str
    nome_db: str
    valor_titular: float
    dependentes: List[DependentConfirmar]
    valor_total: float
    centro_custo: Optional[str] = "N/D"

class ConfirmarImportacaoSorrisoPayload(BaseModel):
    nomeArquivo: str
    titulares: List[TitularConfirmar]
    idEmpresa: Optional[int] = None

@router.post("/plano-saude/sorriso/confirmar")
def confirmar_importacao_plano_saude_sorriso(
    payload: ConfirmarImportacaoSorrisoPayload,
    db: Session = Depends(get_db)
):
    try:
        from app.models.importacao import Importacao
        from app.models.movimentacao import Movimentacao
        from app.models.empresa import Empresa
        from app.repositories.empresa_repository import EmpresaRepository
        from app.repositories.colaborador_repository import ColaboradorRepository
        from app.repositories.categoria_repository import CategoriaRepository
        from app.schemas.categoria import CategoriaCreate
        
        emp_repo = EmpresaRepository(db)
        colab_repo = ColaboradorRepository(db)
        cat_repo = CategoriaRepository(db)
        
        emp = None
        if payload.idEmpresa is not None:
            emp = emp_repo.get_by_id(payload.idEmpresa)
            
        if not emp:
            emp = emp_repo.get_by_nome("RDV - SANTA MARIA")
        if not emp:
            from app.models.empresa import Empresa as ModelEmpresa
            empresas_todas = db.query(ModelEmpresa).all()
            if empresas_todas:
                emp = empresas_todas[0]
            else:
                raise HTTPException(status_code=400, detail="Empresa RDV - SANTA MARIA não encontrada.")
                
        is_seguro = "seguro" in emp.nome.lower()
        
        if is_seguro:
            cat = cat_repo.get_by_id(9)
            if not cat:
                cat = cat_repo.get_by_nome("Seguro/Saúde")
            if not cat:
                cat = cat_repo.get_by_nome("Seguro/Saude")
            if not cat:
                try:
                    novo_cat = CategoriaCreate(nome="Seguro/Saúde", descricao="Despesas com seguros e saúde")
                    cat = cat_repo.create(novo_cat)
                except Exception:
                    cat = None
            if not cat:
                class MockCat:
                    idCategorias = 9
                cat = MockCat()
        else:
            cat = cat_repo.get_by_nome("Plano de Saúde")
            if not cat:
                cat = cat_repo.get_by_nome("Plano de Saude")
            if not cat:
                try:
                    novo_cat = CategoriaCreate(nome="Plano de Saúde", descricao="Despesas com planos de saúde e odontológicos")
                    cat = cat_repo.create(novo_cat)
                except Exception:
                    cat = cat_repo.get_by_id(8)
                    if not cat:
                        raise HTTPException(status_code=400, detail="Categoria Plano de Saúde não encontrada e fallback falhou.")
                        
        extensao = payload.nomeArquivo.split('.')[-1] if '.' in payload.nomeArquivo else 'pdf'
        tipo_importacao = "SEGURO" if is_seguro else "PLANO_SAUDE"
        
        nova_importacao = Importacao(
            nomeArquivo=payload.nomeArquivo,
            extensaoArquivo=extensao,
            idEmpresa=emp.idEmpresas,
            tipo=tipo_importacao
        )
        db.add(nova_importacao)
        db.flush() # Gerar idImportacoes
        
        movimentacoes_criadas = 0
        erros_colaboradores = []
        
        # 4. Iterar titulares e criar movimentações
        for t in payload.titulares:
            colab = colab_repo.get_by_nome(t.nome_db)
            if not colab:
                colab = colab_repo.get_by_nome(t.nome_pdf)
            if not colab:
                from app.models.colaborador import Colaborador as ModelColab
                colab = db.query(ModelColab).filter(ModelColab.nome.ilike(t.nome_db)).first()
            if not colab:
                clean_name = t.nome_db.replace(" da ", " ").replace(" de ", " ").replace(" dos ", " ").replace(" do ", " ").replace(" e ", " ")
                colab = db.query(ModelColab).filter(ModelColab.nome.ilike(f"%{clean_name}%")).first()
                
            if not colab:
                erros_colaboradores.append(t.nome_db)
                continue
                
            # --- SALVAR O ALIAS / APRENDIZADO ---
            if t.nome_pdf and t.nome_pdf.strip().upper() != colab.nome.strip().upper():
                try:
                    from app.repositories.colaborador_alias_repository import ColaboradorAliasRepository
                    alias_repo = ColaboradorAliasRepository(db)
                    alias_repo.create_or_update(colab.idColaborador, t.nome_pdf.strip())
                except Exception as ex:
                    print(f"[WARN] Falha ao salvar Alias de colaborador: {ex}")
                
            # Atualizar Centro de Custo do Colaborador se foi modificado
            if t.centro_custo and t.centro_custo != "N/D":
                try:
                    cc_code = int(t.centro_custo.strip())
                    from app.repositories.centro_custo_repository import centro_custo_repository
                    cc_db = centro_custo_repository.get_by_codigo(db, cc_code)
                    if cc_db and colab.idCentroCusto != cc_db.idCentroCusto:
                        colab.idCentroCusto = cc_db.idCentroCusto
                        db.add(colab)
                except Exception as ex:
                    print(f"[WARN] Falha ao atualizar Centro de Custo do Colaborador: {ex}")
                
            nova_mov = Movimentacao(
                idCategoria=cat.idCategorias,
                idColaborador=colab.idColaborador,
                idEmpresa=emp.idEmpresas,
                idImportacoes=nova_importacao.idImportacoes,
                valor=t.valor_total
            )
            db.add(nova_mov)
            movimentacoes_criadas += 1
            
        if erros_colaboradores:
            print(f"[WARN] Colaboradores não encontrados: {erros_colaboradores}")
            
        db.commit()
        return {
            "sucesso": True,
            "idImportacoes": nova_importacao.idImportacoes,
            "movimentacoes_criadas": movimentacoes_criadas,
            "erros_colaboradores": erros_colaboradores
        }
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


class ExportarSorrisoExcelPayload(BaseModel):
    titulares: List[TitularConfirmar]

@router.post("/plano-saude/sorriso/exportar")
def exportar_sorriso_excel(
    payload: ExportarSorrisoExcelPayload
):
    try:
        rows = []
        total_geral = 0.0
        for t in payload.titulares:
            rows.append({
                "Beneficiário (Titular)": t.nome_db or t.nome_pdf,
                "Centro de Custo": t.centro_custo or "N/D",
                "Valor Total": t.valor_total
            })
            total_geral += t.valor_total
            
        # Linha do Total Geral
        rows.append({
            "Beneficiário (Titular)": "TOTAL GERAL",
            "Centro de Custo": "",
            "Valor Total": total_geral
        })
        
        df = pd.DataFrame(rows)
        
        # Formatar valor total como R$ string
        df["Valor Total"] = df["Valor Total"].apply(lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Consolidação Sorriso')
            
        output.seek(0)
        
        headers_response = {
            'Content-Disposition': 'attachment; filename="planilha_consolidada_sorriso.xlsx"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_response
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/plano-saude/unimed-odonto/analisar")
async def analisar_plano_saude_unimed_odonto(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    try:
        from app.models.colaborador import Colaborador
        from app.repositories.colaborador_repository import ColaboradorRepository
        from sqlalchemy.orm import joinedload
        
        colab_repo = ColaboradorRepository(db)
        content = await file.read()
        
        colabs_db = db.query(Colaborador).all()
        nomes_colaboradores = [c.nome for c in colabs_db]
        
        ia = IAService()
        res_ia = await ia.analisar_plano_saude_unimed_odonto(
            file_content=content,
            file_name=file.filename,
            colaboradores=nomes_colaboradores
        )
        
        titulares_extraidos = res_ia.get("titulares", [])
        
        for t in titulares_extraidos:
            colab = db.query(Colaborador).options(
                joinedload(Colaborador.centro_custo),
                joinedload(Colaborador.unidade)
            ).filter(Colaborador.nome == t.get("nome_db", "")).first()
            
            if not colab:
                colab = db.query(Colaborador).options(
                    joinedload(Colaborador.centro_custo),
                    joinedload(Colaborador.unidade)
                ).filter(Colaborador.nome == t.get("nome_pdf", "")).first()
                
            if colab:
                if colab.centro_custo:
                    t["centro_custo"] = str(colab.centro_custo.codigo)
                else:
                    t["centro_custo"] = "N/D"
                if colab.unidade:
                    t["unidade"] = str(colab.unidade.codigo)
                else:
                    t["unidade"] = "N/D"
            else:
                t["centro_custo"] = "N/D"
                t["unidade"] = "N/D"
        
        nomes_titulares = [t["nome_pdf"].strip().upper() for t in titulares_extraidos]
        titulares_unicos = set(nomes_titulares)
        
        nomes_dependentes = []
        for t in titulares_extraidos:
            for d in t.get("dependentes", []):
                nomes_dependentes.append(d["nome"].strip().upper())
                
        intersection = titulares_unicos.intersection(set(nomes_dependentes))
        
        soma_individual = 0.0
        soma_grupo = 0.0
        for t in titulares_extraidos:
            val_tit = t["valor_titular"]
            val_deps = sum(d["valor"] for d in t.get("dependentes", []))
            soma_individual += val_tit + val_deps
            soma_grupo += t["valor_total"]
            
        validacoes = {
            "apenas_titulares_na_tabela": True,
            "sem_titulares_duplicados": len(nomes_titulares) == len(titulares_unicos),
            "sem_dependentes_como_titulares": len(intersection) == 0,
            "soma_individual_bate_com_total_geral": round(soma_individual, 2) == round(soma_grupo, 2),
            "titulares_count": len(titulares_extraidos),
            "dependentes_count": len(nomes_dependentes),
            "total_count": len(titulares_extraidos) + len(nomes_dependentes)
        }
        
        validacoes_sucesso = all([
            validacoes["sem_titulares_duplicados"],
            validacoes["sem_dependentes_como_titulares"],
            validacoes["soma_individual_bate_com_total_geral"]
        ])
        
        return {
            "sucesso": True,
            "dados": titulares_extraidos,
            "validacoes": validacoes,
            "validacoes_sucesso": validacoes_sucesso,
            "total_geral": round(soma_grupo, 2)
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


class DependentConfirmarUnimed(BaseModel):
    nome: str
    tipo: str
    valor: float

class TitularConfirmarUnimed(BaseModel):
    nome_pdf: str
    nome_db: str
    matricula: str
    valor_titular: float
    dependentes: List[DependentConfirmarUnimed]
    valor_total: float
    centro_custo: Optional[str] = "N/D"
    unidade: Optional[str] = "N/D"

class ConfirmarImportacaoUnimedPayload(BaseModel):
    nomeArquivo: str
    titulares: List[TitularConfirmarUnimed]
    idEmpresa: Optional[int] = None

@router.post("/plano-saude/unimed-odonto/confirmar")
def confirmar_importacao_plano_saude_unimed_odonto(
    payload: ConfirmarImportacaoUnimedPayload,
    db: Session = Depends(get_db)
):
    try:
        from app.models.importacao import Importacao
        from app.models.movimentacao import Movimentacao
        from app.models.empresa import Empresa
        from app.repositories.empresa_repository import EmpresaRepository
        from app.repositories.colaborador_repository import ColaboradorRepository
        from app.repositories.categoria_repository import CategoriaRepository
        from app.schemas.categoria import CategoriaCreate
        
        emp_repo = EmpresaRepository(db)
        colab_repo = ColaboradorRepository(db)
        cat_repo = CategoriaRepository(db)
        
        emp = None
        if payload.idEmpresa is not None:
            emp = emp_repo.get_by_id(payload.idEmpresa)
            
        if not emp:
            emp = emp_repo.get_by_nome("RDV - SANTA MARIA")
        if not emp:
            from app.models.empresa import Empresa as ModelEmpresa
            empresas_todas = db.query(ModelEmpresa).all()
            if empresas_todas:
                emp = empresas_todas[0]
            else:
                raise HTTPException(status_code=400, detail="Empresa RDV - SANTA MARIA não encontrada.")
                
        is_seguro = "seguro" in emp.nome.lower()
        
        if is_seguro:
            cat = cat_repo.get_by_id(9)
            if not cat:
                cat = cat_repo.get_by_nome("Seguro/Saúde")
            if not cat:
                cat = cat_repo.get_by_nome("Seguro/Saude")
            if not cat:
                try:
                    novo_cat = CategoriaCreate(nome="Seguro/Saúde", descricao="Despesas com seguros e saúde")
                    cat = cat_repo.create(novo_cat)
                except Exception:
                    cat = None
            if not cat:
                class MockCat:
                    idCategorias = 9
                cat = MockCat()
        else:
            cat = cat_repo.get_by_nome("Plano de Saúde")
            if not cat:
                cat = cat_repo.get_by_nome("Plano de Saude")
            if not cat:
                try:
                    novo_cat = CategoriaCreate(nome="Plano de Saúde", descricao="Despesas com planos de saúde e odontológicos")
                    cat = cat_repo.create(novo_cat)
                except Exception:
                    cat = cat_repo.get_by_id(8)
                    if not cat:
                        raise HTTPException(status_code=400, detail="Categoria Plano de Saúde não encontrada e fallback falhou.")
                        
        extensao = payload.nomeArquivo.split('.')[-1] if '.' in payload.nomeArquivo else 'pdf'
        tipo_importacao = "SEGURO" if is_seguro else "PLANO_SAUDE"
        
        nova_importacao = Importacao(
            nomeArquivo=payload.nomeArquivo,
            extensaoArquivo=extensao,
            idEmpresa=emp.idEmpresas,
            tipo=tipo_importacao
        )
        db.add(nova_importacao)
        db.flush()
        
        movimentacoes_criadas = 0
        erros_colaboradores = []
        
        for t in payload.titulares:
            colab = colab_repo.get_by_nome(t.nome_db)
            if not colab:
                colab = colab_repo.get_by_nome(t.nome_pdf)
            if not colab:
                from app.models.colaborador import Colaborador as ModelColab
                colab = db.query(ModelColab).filter(ModelColab.nome.ilike(t.nome_db)).first()
            if not colab:
                clean_name = t.nome_db.replace(" da ", " ").replace(" de ", " ").replace(" dos ", " ").replace(" do ", " ").replace(" e ", " ")
                colab = db.query(ModelColab).filter(ModelColab.nome.ilike(f"%{clean_name}%")).first()
                
            if not colab:
                erros_colaboradores.append(t.nome_db)
                continue
                
            if t.centro_custo and t.centro_custo != "N/D":
                try:
                    cc_code = int(t.centro_custo.strip())
                    from app.repositories.centro_custo_repository import centro_custo_repository
                    cc_db = centro_custo_repository.get_by_codigo(db, cc_code)
                    if cc_db and colab.idCentroCusto != cc_db.idCentroCusto:
                        colab.idCentroCusto = cc_db.idCentroCusto
                        db.add(colab)
                except Exception as ex:
                    print(f"[WARN] Falha ao atualizar Centro de Custo do Colaborador: {ex}")
                    
            if t.unidade and t.unidade != "N/D":
                try:
                    from app.models.unidade import Unidade as ModelUnidade
                    unidade_db = db.query(ModelUnidade).filter(ModelUnidade.codigo == int(t.unidade.strip())).first()
                    if unidade_db and colab.idUnidade != unidade_db.idUnidade:
                        colab.idUnidade = unidade_db.idUnidade
                        db.add(colab)
                except Exception as ex:
                    print(f"[WARN] Falha ao atualizar Unidade do Colaborador: {ex}")
                
            nova_mov = Movimentacao(
                idCategoria=cat.idCategorias,
                idColaborador=colab.idColaborador,
                idEmpresa=emp.idEmpresas,
                idImportacoes=nova_importacao.idImportacoes,
                valor=t.valor_total
            )
            db.add(nova_mov)
            movimentacoes_criadas += 1
            
        if erros_colaboradores:
            print(f"[WARN] Colaboradores não encontrados: {erros_colaboradores}")
            
        db.commit()
        return {
            "sucesso": True,
            "idImportacoes": nova_importacao.idImportacoes,
            "movimentacoes_criadas": movimentacoes_criadas,
            "erros_colaboradores": erros_colaboradores
        }
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


class ExportarUnimedExcelPayload(BaseModel):
    titulares: List[TitularConfirmarUnimed]

@router.post("/plano-saude/unimed-odonto/exportar")
def exportar_unimed_odonto_excel(
    payload: ExportarUnimedExcelPayload
):
    try:
        rows = []
        total_geral = 0.0
        for t in payload.titulares:
            rows.append({
                "Unidade": t.unidade or "N/D",
                "Beneficiário (Titular)": t.nome_db or t.nome_pdf,
                "Centro de Custo": t.centro_custo or "N/D",
                "Valor Total": t.valor_total
            })
            total_geral += t.valor_total
            
        rows.append({
            "Unidade": "",
            "Beneficiário (Titular)": "TOTAL GERAL",
            "Centro de Custo": "",
            "Valor Total": total_geral
        })
        
        df = pd.DataFrame(rows)
        df["Valor Total"] = df["Valor Total"].apply(lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Consolidação Unimed')
            
        output.seek(0)
        
        headers_response = {
            'Content-Disposition': 'attachment; filename="planilha_consolidada_unimed_odonto.xlsx"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_response
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))






@router.post("/cema/conciliar")
async def conciliar_cema(
    cema_file: UploadFile = File(...),
    acr_file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not cema_file.filename or not acr_file.filename:
        raise HTTPException(status_code=400, detail="Arquivos inválidos")
        
    try:
        import io
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        import datetime
        import math
        
        # 1. Parse Cema File
        cema_bytes = await cema_file.read()
        try:
            df_cema = pd.read_excel(io.BytesIO(cema_bytes), header=None)
        except Exception:
            raise HTTPException(status_code=400, detail="Arquivo Cema deve ser Excel.")
            
        # Parse Dates Helper
        def parse_date(date_str):
            if not date_str:
                return None
            if isinstance(date_str, datetime.datetime):
                return date_str.date()
            if isinstance(date_str, datetime.date):
                return date_str
            date_str = str(date_str).strip()
            if hasattr(pd, 'Timestamp') and isinstance(date_str, pd.Timestamp):
                return date_str.date()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%y', '%Y/%m/%d %H:%M:%S', '%d/%m/%Y %H:%M:%S'):
                try:
                    return datetime.datetime.strptime(date_str.split()[0] if ' ' in date_str else date_str, fmt).date()
                except ValueError:
                    continue
            return None

        # Next Wednesday Helper
        def get_next_wednesday(d):
            if not isinstance(d, datetime.date):
                return d
            wd = d.weekday()
            if wd == 2:
                return d
            days_ahead = 2 - wd
            if days_ahead <= 0:
                days_ahead += 7
            return d + datetime.timedelta(days_ahead)

        cema_invoices = []
        for idx, row in df_cema.iterrows():
            if len(row) < 6:
                continue
            
            nf_raw = str(row[1]).strip() # Col B (index 1)
            date_raw = row[5] # Col F (index 5)
            
            if not nf_raw or nf_raw.lower() in ('nan', 'none', 'nota fiscal', 'nota', 'nota_fiscal'):
                continue
                
            nf_clean = nf_raw.split('/')[0].strip()
            
            if not nf_clean:
                continue
                
            if not any(char.isdigit() for char in nf_clean):
                continue
            
            try:
                nf_clean = str(int(float(nf_clean))).zfill(7)
            except ValueError:
                nf_clean = nf_clean.zfill(7)
            
            d_parsed = parse_date(date_raw)
            d_adjusted = get_next_wednesday(d_parsed) if d_parsed else date_raw
            
            cema_invoices.append({
                'raw_nf': nf_raw,
                'nf': nf_clean,
                'prorrogacao': d_adjusted
            })
            
        if not cema_invoices:
            raise HTTPException(status_code=400, detail="Nenhuma nota fiscal encontrada no arquivo Cema.")

        # 2. Parse ACR File
        acr_bytes = await acr_file.read()
        try:
            df_acr = pd.read_excel(io.BytesIO(acr_bytes), header=None)
        except Exception:
            acr_text = acr_bytes.decode('utf-8', errors='ignore')
            sep = ';' if ';' in acr_text else ','
            df_acr = pd.read_csv(io.StringIO(acr_text), sep=sep, header=None)
            
        acr_data_by_int = {}
        for idx, row in df_acr.iterrows():
            if len(row) < 16:
                continue
                
            # For Cema, just extract Nota (Col D / index 3) and Vencimento (Col P / index 15)
            # Remove strict 104, DP, and Parcela filters which could cause "Não encontrado"
            val_d_raw = str(row[3]).strip()
            if not val_d_raw or val_d_raw.lower() in ('nan', 'none'):
                continue
                
            val_d = val_d_raw.split('.')[0]
            val_p = str(row[15]).strip()
            if ' ' in val_p:
                val_p = val_p.split()[0]
                
            try:
                acr_data_by_int[int(val_d)] = parse_date(val_p) or val_p
            except ValueError:
                continue
                
        # 3. Create Excel
        def format_date_to_br(d):
            if isinstance(d, datetime.date):
                return d.strftime('%d/%m/%Y')
            return str(d) if d else ""

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Prorrogações Cema"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.cell(row=1, column=1, value="Nota Fiscal")
        ws1.cell(row=1, column=2, value="Data de Prorrogação")
        
        for i, inv in enumerate(cema_invoices):
            ws1.cell(row=i+2, column=1, value=inv['nf'])
            ws1.cell(row=i+2, column=2, value=format_date_to_br(inv['prorrogacao']))
            
        ws2 = wb.create_sheet(title="Conciliação")
        ws2.views.sheetView[0].showGridLines = True
        ws2.cell(row=1, column=1, value="Nota Fiscal")
        ws2.cell(row=1, column=2, value="Data Cema (Prorrogação)")
        ws2.cell(row=1, column=3, value="Data ACR (Vencimento)")
        ws2.cell(row=1, column=4, value="Status")
        
        font_header = Font(name='Calibri', size=11, bold=True)
        font_body = Font(name='Calibri', size=11, bold=False)
        align_left = Alignment(horizontal='left', vertical='center')
        align_center = Alignment(horizontal='center', vertical='center')
        
        fill_ok = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fill_div = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        for col_idx in [1, 2]:
            cell = ws1.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left
            
        for r_idx in range(2, len(cema_invoices) + 2):
            ws1.cell(row=r_idx, column=1).font = font_body
            ws1.cell(row=r_idx, column=1).number_format = '@'
            ws1.cell(row=r_idx, column=1).alignment = align_left
            ws1.cell(row=r_idx, column=2).font = font_body
            ws1.cell(row=r_idx, column=2).alignment = align_center
            
        ws1.column_dimensions['A'].width = 16
        ws1.column_dimensions['B'].width = 24
        
        for i, inv in enumerate(cema_invoices):
            row_num = i + 2
            nf_str = inv['nf']
            d_cema = inv['prorrogacao']
            
            try:
                nf_int = int(nf_str)
            except ValueError:
                nf_int = None
                
            d_acr = None
            status = "Não encontrado no CSV"
            status_fill = None
            
            if nf_int is not None and nf_int in acr_data_by_int:
                d_acr = acr_data_by_int[nf_int]
                
                if isinstance(d_cema, datetime.date) and isinstance(d_acr, datetime.date):
                    if d_cema == d_acr:
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                else:
                    if str(d_cema).strip() == str(d_acr).strip():
                        status = "OK"
                        status_fill = fill_ok
                    else:
                        status = "Divergente"
                        status_fill = fill_div
                        
            ws2.cell(row=row_num, column=1, value=nf_str)
            ws2.cell(row=row_num, column=2, value=format_date_to_br(d_cema))
            ws2.cell(row=row_num, column=3, value=format_date_to_br(d_acr))
            
            status_cell = ws2.cell(row=row_num, column=4, value=status)
            if status_fill:
                status_cell.fill = status_fill
                
        for col_idx in [1, 2, 3, 4]:
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.alignment = align_left
            
        for r_idx in range(2, len(cema_invoices) + 2):
            ws2.cell(row=r_idx, column=1).font = font_body
            ws2.cell(row=r_idx, column=1).number_format = '@'
            ws2.cell(row=r_idx, column=1).alignment = align_left
            ws2.cell(row=r_idx, column=2).font = font_body
            ws2.cell(row=r_idx, column=2).alignment = align_center
            ws2.cell(row=r_idx, column=3).font = font_body
            ws2.cell(row=r_idx, column=3).alignment = align_center
            ws2.cell(row=r_idx, column=4).font = font_body
            ws2.cell(row=r_idx, column=4).alignment = align_left
            
        ws2.column_dimensions['A'].width = 16
        ws2.column_dimensions['B'].width = 24
        ws2.column_dimensions['C'].width = 24
        ws2.column_dimensions['D'].width = 24
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = cema_file.filename.rsplit('.', 1)[0] + "_extraido.xlsx"
        
        from app.services.importacao_service import ImportacaoService
        ImportacaoService(db).registrar_importacao(filename, "xlsx", "Prorrogação - Cema")
        
        headers_response = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_response
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
