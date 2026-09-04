import io
import re
from typing import Dict, List, Optional
from sqlalchemy.orm import Session
import openpyxl
from fastapi import HTTPException

from app.models.unidade import Unidade
from app.models.colaborador import Colaborador
from app.repositories.colaborador_repository import ColaboradorRepository
from app.repositories.centro_custo_repository import centro_custo_repository
from app.repositories.unidade_repository import unidade_repository
from app.repositories.cargo_colaborador_repository import CargoColaboradorRepository
from app.schemas.colaborador_import import (
    ImportUnidade, ImportNovo, ImportDivergente, ImportDesligado, ImportErro,
    ImportPreviewResponse, ImportProcessarRequest, ImportProcessarResponse
)

HEADER_ROW = 2
DATA_START_ROW = 3


def _normalizar_documento(valor) -> Optional[str]:
    if valor is None:
        return None
    if isinstance(valor, float):
        valor = int(valor)
    digitos = re.sub(r"\D", "", str(valor))
    if not digitos:
        return None
    if len(digitos) < 11:
        digitos = digitos.zfill(11)
    return digitos


def _normalizar_nome(valor) -> Optional[str]:
    if valor is None:
        return None
    texto = " ".join(str(valor).split())
    return texto.title() if texto else None


def _normalizar_codigo(valor) -> Optional[int]:
    if valor is None or str(valor).strip() == "":
        return None
    try:
        return int(float(str(valor).strip()))
    except ValueError:
        return None


def _to_import_unidade(unidade: Unidade) -> ImportUnidade:
    return ImportUnidade(idUnidade=unidade.idUnidade, codigo=unidade.codigo, descricao=unidade.descricao)


class ColaboradorImportService:
    def __init__(self, db: Session):
        self.db = db
        self.colab_repo = ColaboradorRepository(db)
        self.cargo_repo = CargoColaboradorRepository(db)

    def _ler_planilha(self, conteudo: bytes):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Não foi possível ler o arquivo Excel: {e}")

        agregados: Dict[str, dict] = {}
        erros: List[ImportErro] = []

        for aba in wb.sheetnames:
            ws = wb[aba]
            for idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, max_col=4, values_only=True), start=DATA_START_ROW):
                estab_raw, nome_raw, doc_raw, cc_raw = (row + (None, None, None, None))[:4]

                nome = _normalizar_nome(nome_raw)
                if not nome:
                    continue

                documento = _normalizar_documento(doc_raw)
                if not documento:
                    erros.append(ImportErro(aba=aba, linha=idx, nome=nome, documento=None,
                                             motivo="Documento não informado ou inválido"))
                    continue

                estab_codigo = _normalizar_codigo(estab_raw)
                cc_codigo = _normalizar_codigo(cc_raw)
                if cc_codigo is None:
                    erros.append(ImportErro(aba=aba, linha=idx, nome=nome, documento=documento,
                                             motivo="Centro de Custo não informado ou inválido"))
                    continue

                unidade = unidade_repository.get_by_codigo(self.db, estab_codigo) if estab_codigo is not None else None
                if not unidade:
                    erros.append(ImportErro(aba=aba, linha=idx, nome=nome, documento=documento,
                                             motivo=f"Unidade/Estab '{estab_raw}' não cadastrada no sistema"))
                    continue

                entrada = agregados.setdefault(documento, {"nome": nome, "unidades": {}, "cc_codigos": set()})
                entrada["nome"] = nome
                entrada["unidades"][unidade.idUnidade] = unidade
                entrada["cc_codigos"].add(cc_codigo)

        return agregados, erros

    def preview(self, conteudo: bytes) -> ImportPreviewResponse:
        agregados, erros = self._ler_planilha(conteudo)

        novos: List[ImportNovo] = []
        divergentes: List[ImportDivergente] = []

        for documento, dados in agregados.items():
            if len(dados["cc_codigos"]) > 1:
                erros.append(ImportErro(
                    aba="-", linha=0, nome=dados["nome"], documento=documento,
                    motivo=f"Centro de Custo divergente para o mesmo documento dentro da planilha: {sorted(dados['cc_codigos'])}"
                ))
                continue

            cc_codigo = next(iter(dados["cc_codigos"]))
            cc = centro_custo_repository.get_by_codigo(self.db, cc_codigo)
            unidades_planilha = list(dados["unidades"].values())
            unidades_import = [_to_import_unidade(u) for u in unidades_planilha]

            existente = self.colab_repo.get_by_documento(documento)

            if not existente:
                novos.append(ImportNovo(
                    documento=documento, nome=dados["nome"], unidades=unidades_import,
                    centroCustoCodigo=cc_codigo,
                    idCentroCusto=cc.idCentroCusto if cc else None,
                    centroCustoNome=cc.nome if cc else None,
                    ccEncontrado=bool(cc)
                ))
                continue

            ids_unidades_planilha = {u.idUnidade for u in unidades_planilha}
            ids_unidades_atuais = {u.idUnidade for u in existente.unidades}
            unidades_divergentes = ids_unidades_planilha != ids_unidades_atuais
            cc_divergente = bool(cc) and existente.idCentroCusto != cc.idCentroCusto
            reativado = existente.snAtivo == 'N'

            precisa_atencao = (not cc) or cc_divergente or unidades_divergentes or reativado
            if not precisa_atencao:
                continue

            divergentes.append(ImportDivergente(
                idColaborador=existente.idColaborador, documento=documento, nome=dados["nome"],
                unidades=unidades_import,
                unidadesAtuais=[_to_import_unidade(u) for u in existente.unidades],
                centroCustoCodigo=cc_codigo,
                idCentroCusto=cc.idCentroCusto if cc else None,
                centroCustoNome=cc.nome if cc else None,
                ccEncontrado=bool(cc),
                idCentroCustoAtual=existente.idCentroCusto,
                centroCustoAtualNome=existente.centro_custo.nome if existente.centro_custo else None,
                ccDivergente=cc_divergente,
                unidadesDivergentes=unidades_divergentes,
                reativado=reativado
            ))

        documentos_planilha = set(agregados.keys())
        desligados: List[ImportDesligado] = []
        for documento in self.colab_repo.get_documentos_ativos():
            if documento in documentos_planilha:
                continue
            colab = self.colab_repo.get_by_documento(documento)
            if not colab:
                continue
            desligados.append(ImportDesligado(
                idColaborador=colab.idColaborador, documento=documento, nome=colab.nome,
                unidadesAtuais=[_to_import_unidade(u) for u in colab.unidades],
                centroCustoAtualNome=colab.centro_custo.nome if colab.centro_custo else None
            ))

        return ImportPreviewResponse(novos=novos, divergentes=divergentes, desligados=desligados, erros=erros)

    def _get_cargo_padrao_id(self) -> int:
        cargo = self.cargo_repo.get_by_nome("Padrão")
        if not cargo:
            itens, total = self.cargo_repo.get_all(limit=1)
            cargo = itens[0] if itens else None
        if not cargo:
            raise HTTPException(status_code=400, detail="Nenhum tipo de colaborador cadastrado para usar como padrão.")
        return cargo.idCargoColaborador

    def processar(self, payload: ImportProcessarRequest) -> ImportProcessarResponse:
        id_cargo_padrao = self._get_cargo_padrao_id()
        cadastrados = 0
        atualizados = 0
        desligados = 0

        try:
            for novo in payload.novos:
                db_obj = Colaborador(
                    nome=novo.nome,
                    documento=novo.documento,
                    idCentroCusto=novo.idCentroCusto,
                    idCargoColaborador=id_cargo_padrao,
                    snAtivo='S'
                )
                self.db.add(db_obj)
                self.db.flush()
                self.colab_repo.sync_unidades(db_obj, novo.unidadeIds)
                cadastrados += 1

            for div in payload.divergentes:
                db_obj = self.db.query(Colaborador).filter(Colaborador.idColaborador == div.idColaborador).first()
                if not db_obj:
                    raise HTTPException(status_code=404, detail=f"Colaborador {div.idColaborador} não encontrado.")
                db_obj.idCentroCusto = div.idCentroCusto
                db_obj.snAtivo = 'S'
                self.colab_repo.sync_unidades(db_obj, div.unidadeIds)
                atualizados += 1

            for desl in payload.desligados:
                db_obj = self.db.query(Colaborador).filter(Colaborador.idColaborador == desl.idColaborador).first()
                if not db_obj:
                    raise HTTPException(status_code=404, detail=f"Colaborador {desl.idColaborador} não encontrado.")
                db_obj.snAtivo = 'N'
                desligados += 1

            self.db.commit()
        except HTTPException:
            self.db.rollback()
            raise
        except Exception as e:
            self.db.rollback()
            raise HTTPException(status_code=500, detail=f"Erro ao processar importação: {e}")

        return ImportProcessarResponse(cadastrados=cadastrados, atualizados=atualizados, desligados=desligados)
