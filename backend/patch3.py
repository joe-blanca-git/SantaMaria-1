import sys
import re

with open('c:/Users/tania.canedo/Documents/git/SantaMaria/backend/app/routers/importacoes.py', 'r', encoding='utf-8') as f:
    content = f.read()

pattern = re.compile(r'async def conciliar_composicao_ws.*?ws2\.column_dimensions\[\'F\'\]\.width = 16\n', re.DOTALL)

new_func = '''async def conciliar_composicao_ws(wb, acr_file, rows_to_export, font_header, font_body, align_center, align_left, align_right):
    if not acr_file or not acr_file.filename:
        return
    acr_bytes = await acr_file.read()
    import io
    import pandas as pd
    from openpyxl.styles import PatternFill
    try:
        df_acr = pd.read_excel(io.BytesIO(acr_bytes), header=None)
    except Exception:
        acr_text = acr_bytes.decode('utf-8', errors='ignore')
        sep = ';' if ';' in acr_text else ','
        df_acr = pd.read_csv(io.StringIO(acr_text), sep=sep, header=None)
        
    col_titulo = 3 # Coluna D
    col_parcela = 4 # Coluna E
    col_saldo = -1
    
    # Busca 1: Pelo nome no cabeçalho
    for idx, row in df_acr.head(50).iterrows():
        row_strs = [str(x).strip().lower() for x in row.values]
        for c_idx, val in enumerate(row_strs):
            if val == 'saldo' or val == 'saldo ' or 'saldo' in val or 'valor l' in val:
                col_saldo = c_idx
                break
        if col_saldo != -1:
            break
            
    # Busca 2 (Heurística Robusta): Encontrar a coluna cruzando com os valores esperados da composição
    if col_saldo == -1:
        comp_map = {}
        for item in rows_to_export:
            nf = str(item.get('Nota Fiscal') or item.get('nf')).strip()
            if nf:
                v = abs(item.get('Valor Liquido') or item.get('valor_liquido') or 0.0)
                if v > 0:
                    comp_map[nf.lstrip('0')] = v
                    
        for idx, row in df_acr.iterrows():
            if len(row) <= col_titulo:
                continue
            val_d_raw = row[col_titulo]
            if pd.isna(val_d_raw):
                continue
            titulo = str(val_d_raw).strip().split('.')[0]
            if titulo.endswith('.0'): titulo = titulo[:-2]
            t_clean = titulo.lstrip('0')
            
            if t_clean in comp_map:
                expected_val = comp_map[t_clean]
                for c_idx, cell in enumerate(row.values):
                    try:
                        if isinstance(cell, str):
                            cell_clean = cell.replace('R$', '').replace('.', '').replace(',', '.').strip()
                            cell_val = abs(float(cell_clean))
                        else:
                            cell_val = abs(float(cell))
                        if cell_val > 0 and abs(cell_val - expected_val) < 0.01:
                            col_saldo = c_idx
                            break
                    except Exception:
                        pass
                if col_saldo != -1:
                    break
                    
    # Fallback final se nada encontrar
    if col_saldo == -1:
        col_saldo = df_acr.shape[1] - 1 if df_acr.shape[1] > 0 else 0

    acr_data = {}
    for idx, row in df_acr.iterrows():
        # Verificação segura de tamanho
        if len(row) <= col_titulo or len(row) <= col_saldo:
            continue
            
        val_d_raw = row[col_titulo]
        if pd.isna(val_d_raw):
            continue
            
        titulo = str(val_d_raw).strip().split('.')[0]
        if not any(char.isdigit() for char in titulo):
            continue
            
        if titulo.endswith('.0'):
            titulo = titulo[:-2]
            
        if col_parcela < len(row):
            parcela_val = str(row[col_parcela]).strip().split('.')[0]
            try:
                if int(parcela_val) != 1:
                    continue
            except ValueError:
                pass
                
        saldo_val = row[col_saldo]
        try:
            if isinstance(saldo_val, str):
                saldo_val = saldo_val.replace('R$', '').replace('.', '').replace(',', '.').strip()
            saldo_float = float(saldo_val)
            # Salvar no mapa pelo lstrip para match infalível
            acr_data[titulo.lstrip('0')] = abs(saldo_float)
        except ValueError:
            continue
            
    ws2 = wb.create_sheet(title="Conciliação")
    headers2 = ['Nota Fiscal', 'Parcela', 'Status NF', 'Valor Composição', 'Valor ACR', 'Status Valor']
    ws2.append(headers2)
    for col_idx, col_name in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.alignment = align_left if col_idx <= 3 else align_right
    fill_ok = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_err = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    accounting_format = '_("R$"* #,##0.00_);_("R$"* (#,##0.00);_("R$"* "-"_);_(@_)'
    
    for i, item in enumerate(rows_to_export):
        nf = item.get('Nota Fiscal') or item.get('nf')
        if not nf:
            continue
        nf_str = str(nf).strip()
        val_comp = item.get('Valor Liquido') or item.get('valor_liquido') or 0.0
        
        status_nf = 'Não Encontrado'
        status_val = '-'
        val_acr = 0.0
        
        # Remove trailing T for match attempt if it's a T parcel
        nf_clean = nf_str.lstrip('0')
        if nf_clean.endswith('T') or nf_clean.endswith('t'):
            nf_base = nf_clean[:-1]
        else:
            nf_base = nf_clean
            
        found_key = None
        if nf_clean in acr_data:
            found_key = nf_clean
        elif nf_base in acr_data:
            found_key = nf_base
                    
        if found_key:
            status_nf = 'Encontrado'
            val_acr = acr_data[found_key]
            if abs(abs(val_comp) - val_acr) < 0.01:
                status_val = 'OK'
            else:
                status_val = 'Divergente'
                
        r_idx = i + 2
        ws2.cell(row=r_idx, column=1, value=nf_str).font = font_body
        ws2.cell(row=r_idx, column=1).alignment = align_left
        ws2.cell(row=r_idx, column=2, value="01").font = font_body
        ws2.cell(row=r_idx, column=2).alignment = align_left
        ws2.cell(row=r_idx, column=3, value=status_nf).font = font_body
        ws2.cell(row=r_idx, column=3).alignment = align_left
        c4 = ws2.cell(row=r_idx, column=4, value=val_comp)
        c4.font = font_body
        c4.number_format = accounting_format
        c4.alignment = align_right
        c5 = ws2.cell(row=r_idx, column=5, value=val_acr if status_nf == 'Encontrado' else None)
        c5.font = font_body
        c5.number_format = accounting_format
        c5.alignment = align_right
        c6 = ws2.cell(row=r_idx, column=6, value=status_val)
        c6.font = font_body
        c6.alignment = align_right
        if status_nf == 'Encontrado' and status_val == 'OK':
            for c in range(1, 7):
                ws2.cell(row=r_idx, column=c).fill = fill_ok
        else:
            for c in range(1, 7):
                ws2.cell(row=r_idx, column=c).fill = fill_err
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 18
    ws2.column_dimensions['F'].width = 16
'''

content = pattern.sub(new_func, content)
with open('c:/Users/tania.canedo/Documents/git/SantaMaria/backend/app/routers/importacoes.py', 'w', encoding='utf-8') as f:
    f.write(content)
print("Updated perfectly")
